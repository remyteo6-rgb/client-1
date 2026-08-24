import os
import time
import re
import json
import unicodedata
import psycopg2
import psycopg2.extras
import openpyxl
from datetime import datetime, timedelta
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from flask import Flask, render_template, request, redirect, url_for, flash, g, abort, session, Response, jsonify
from parser import (
    parse_sportscode_xml, aggregate_match_stats, aggregate_zones, CATEGORY_SECTIONS,
    SECTION_ICONS, SECTION_HELP, CATEGORY_HELP, generate_highlights, compute_radar_metrics,
    compute_score, compute_phase_timing, compute_attack_sector, compute_defense_sector,
    compute_lineout_detail, compute_scrum_detail, compute_kicking_detail,
    compute_player_attack_table, compute_player_defense_table, compute_player_ruck_table,
    compute_overview_dashboard, compute_ruck_sector, compute_season_dashboard,
    compute_squad_preview, compute_squad_season_stats, SQUAD_ROSTER, SQUAD_POSITION_ORDER,
    is_jiff, compute_jiff_chart,
    compute_transition_sector, compute_player_comparison, compute_player_radar_svg,
    compute_back3_trend, TRAINING_TAXONOMY, compute_training_volume,
    group_training_sessions_by_period, compute_win_loss_analysis, compute_match_kpis,
   PHASE_ICONS, PHASE_HELP, compute_event_timing_multi, compute_match_baseline,
    compute_sector_baselines, compute_player_season_baselines, build_player_cards,
    attach_overview_highlights,
)
from parser_ubb import parse_ubb_xml, compute_ubb_overview
from prod2 import (
    parse_prod2_report, get_team_profile, get_classement_table, compute_player_groups,
    SECTOR_SHEETS, compute_team_trends, compute_most_used_players, build_compare_rows,
    compute_team_kpi_profile, compute_team_threats, compute_team_position_history,
)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "data", "uploads")
# Base de données PostgreSQL persistante (Neon, ou toute autre base Postgres gratuite).
# À définir dans les variables d'environnement Render sous le nom DATABASE_URL.
# Exemple de valeur : postgresql://user:motdepasse@hote.neon.tech/rugby?sslmode=require
DATABASE_URL = os.environ.get("DATABASE_URL")
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-me")
app.config["MAX_CONTENT_LENGTH"] = 30 * 1024 * 1024  # 30MB max upload
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=30)  # reste connecté 30 jours
ASSET_VERSION = str(int(time.time()))  # change à chaque redémarrage : force le navigateur à
                                        # retélécharger le CSS/JS au lieu de garder une vieille
                                        # version en cache après un déploiement.
# Identité du club : seules ces 2 variables (+ le fichier static/logo.png) changent d'une
# copie du site à l'autre pour un autre club. CLUB_NAME = nom court utilisé dans le texte des
# pages ("Ce que {{ CLUB_NAME }} a fait..."), CLUB_FULL_NAME = nom complet affiché dans le
# logo/titre du site. Définis-les dans les variables d'environnement Render pour un nouveau
# client plutôt que de les changer dans le code.
CLUB_NAME = os.environ.get("CLUB_NAME", "Nice")
CLUB_FULL_NAME = os.environ.get("CLUB_FULL_NAME", "Nissa Rugby")
# Le module Adversaires/Pro D2 (scouting hebdo à partir du rapport Excel de la ligue) est
# spécifique à la Pro D2 française : à désactiver pour un club qui n'y a pas accès (mets
# ENABLE_PROD2=0 dans les variables d'environnement Render). Le quota JIFF (règle LNR,
# indépendante de la Pro D2 en tant que telle) se désactive séparément avec ENABLE_JIFF=0.
# Les deux restent activés par défaut pour ne rien changer au site actuel.
ENABLE_PROD2 = os.environ.get("ENABLE_PROD2", "1") != "0"
ENABLE_JIFF = os.environ.get("ENABLE_JIFF", "1") != "0"
PROD2_ENDPOINTS = {
    "opponents", "opponent_detail", "opponent_joueurs", "opponents_trends",
    "opponent_sector", "prod2_import", "next_match", "next_match_set",
}
JIFF_ENDPOINTS = {"season_jiff"}
# Identifiants du compte ADMIN (peut tout faire : importer/supprimer un match, modifier
# la composition, sauvegarder/restaurer...). Définis ADMIN_EMAIL / ADMIN_PASSWORD dans les
# variables d'environnement Render — sinon ces valeurs par défaut (à changer !) sont utilisées.
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "admin@nissarugby.fr")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "changeme")
# Lien de démo public : quiconque a ce lien voit le site en lecture seule, sans compte,
# avec le mode démo (noms floutés) activé automatiquement. Change cette valeur si tu veux
# un lien différent, et ne le partage qu'avec des prospects (il donne accès en lecture à tout).
DEMO_TOKEN = os.environ.get("DEMO_TOKEN", "decouverte-club1")
# Comptes STAFF (accès lecture seule : navigue partout mais ne voit aucune action de
# modification) : autant de comptes que voulu, chacun avec SON PROPRE email/mot de passe.
# Sur Render, ajoute des variables d'environnement par paire numérotée :
#   STAFF_EMAIL_1 / STAFF_PASSWORD_1
#   STAFF_EMAIL_2 / STAFF_PASSWORD_2
#   STAFF_EMAIL_3 / STAFF_PASSWORD_3
#   ... (jusqu'à 20 comptes staff possibles)
# L'ancienne paire sans numéro (STAFF_EMAIL / STAFF_PASSWORD), si tu l'avais déjà
# configurée, continue aussi de fonctionner comme un compte staff de plus.
def _load_staff_accounts():
    accounts = []
    legacy_email = os.environ.get("STAFF_EMAIL")
    legacy_password = os.environ.get("STAFF_PASSWORD")
    if legacy_email and legacy_password:
        accounts.append((legacy_email.strip().lower(), legacy_password))
    for i in range(1, 21):
        email = os.environ.get(f"STAFF_EMAIL_{i}")
        password = os.environ.get(f"STAFF_PASSWORD_{i}")
        if email and password:
            accounts.append((email.strip().lower(), password))
    return accounts
STAFF_ACCOUNTS = _load_staff_accounts()
# Le site entier est privé : seule une personne connectée (admin OU staff) peut voir
# quoi que ce soit. Seules ces 2 routes restent accessibles sans connexion (sinon
# impossible d'atteindre la page de connexion elle-même).
PUBLIC_ENDPOINTS = {"login", "static", "demo_login", "pwa_manifest", "pwa_service_worker"}
@app.before_request
def require_login_everywhere():
    if request.endpoint is None or request.endpoint in PUBLIC_ENDPOINTS:
        return
    if not session.get("logged_in"):
        return redirect(url_for("login", next=request.full_path))
@app.before_request
def gate_optional_features():
    """Bloque proprement les pages Adversaires/Pro D2 et JIFF quand elles sont désactivées
    pour ce club (voir ENABLE_PROD2 / ENABLE_JIFF), plutôt que de les laisser planter sur
    l'absence de données qu'elles ne pourront jamais avoir."""
    if not ENABLE_PROD2 and request.endpoint in PROD2_ENDPOINTS:
        flash("Cette fonctionnalité n'est pas activée sur ce site.", "error")
        return redirect(url_for("landing"))
    if not ENABLE_JIFF and request.endpoint in JIFF_ENDPOINTS:
        flash("Cette fonctionnalité n'est pas activée sur ce site.", "error")
        return redirect(url_for("landing"))
    # Les espaces Documents / Calendrier contiennent des informations internes au staff :
    # ils restent totalement invisibles pour les visiteurs du lien de démonstration. Le
    # Cahier des charges (dont le P.P.I.D) est, lui, consultable en démo — pour montrer la
    # fonctionnalité — mais uniquement en lecture : toute action qui modifie des données
    # réelles (tâches, documents, évaluations PPID...) reste bloquée, quel que soit le
    # préfixe de la page qui la déclenche.
    demo_blocked_prefixes = ("documents", "calendrier", "ppid_")
    demo_blocked_actions = {
        "cahier_charges_add", "cahier_charges_status", "cahier_charges_delete",
        "cahier_charges_upload", "cahier_charges_add_link", "cahier_charges_doc_delete",
    }
    if session.get("demo_forced") and request.endpoint and (
        request.endpoint.startswith(demo_blocked_prefixes)
        or request.endpoint in demo_blocked_actions
    ):
        flash("Cet espace n'est pas accessible en mode démonstration.", "error")
        return redirect(url_for("landing"))
# Un compte joueur n'a accès qu'à un petit sous-ensemble du site : son espace, le
# calendrier (en lecture seule — les routes qui ajoutent/modifient/suppriment des
# événements ne sont volontairement PAS dans cette liste), les documents qui lui sont
# partagés, et ses propres stats. Liste blanche plutôt que liste noire : plus sûr,
# ça ne dépend pas de penser à bloquer chaque nouvelle page d'analyse à l'avenir.
PLAYER_ALLOWED_ENDPOINTS = {
    "player_home", "logout", "static", "pwa_manifest", "pwa_service_worker",
    "calendrier", "calendrier_api_events",
    "player_documents", "player_document_download", "player_document_preview",
    "player_stats",
    "player_evaluations", "player_ppid_auto_update",
}
@app.before_request
def gate_player_access():
    if session.get("is_player") and request.endpoint and request.endpoint not in PLAYER_ALLOWED_ENDPOINTS:
        flash("Cette page n'est pas accessible depuis un compte joueur.", "error")
        return redirect(url_for("player_home"))
def admin_required(view):
    """Garde-fou pour les actions réservées à l'admin (import, suppression, validation
    composition, sauvegarde...). Le staff est bien connecté (passe le before_request
    global) mais n'a pas le droit d'exécuter ces actions."""
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("is_admin"):
            flash("Action réservée à l'administrateur.", "error")
            return redirect(url_for("landing"))
        return view(*args, **kwargs)
    return wrapped
@app.context_processor
def inject_logged_in():
    return {
        "logged_in": session.get("logged_in", False), "is_admin": session.get("is_admin", False),
        "is_player": session.get("is_player", False),
        "demo_forced": session.get("demo_forced", False),
        "user_email": session.get("user_email", ""),
        "club_name": CLUB_NAME, "club_full_name": CLUB_FULL_NAME,
        "enable_prod2": ENABLE_PROD2, "enable_jiff": ENABLE_JIFF,
        "asset_version": ASSET_VERSION,
    }

# Contenu du service worker, généré en Python pour pouvoir y injecter ASSET_VERSION :
# le nom du cache change donc à chaque redémarrage/déploiement, ce qui vide
# automatiquement l'ancien cache (voir le handler "activate" ci-dessous) sans
# jamais servir une vieille version du CSS/JS aux téléphones qui ont installé
# l'appli.
SERVICE_WORKER_JS = """
const CACHE_NAME = "rugby-analytics-shell-%(v)s";
const APP_SHELL = [
  "/static/style.css",
  "/static/demo-mode.js",
  "/static/sortable.js",
  "/static/calendar.js",
  "/static/logo.png",
  "/static/icon-192.png",
  "/static/icon-512.png",
];

self.addEventListener("install", (event) => {
  event.waitUntil(
    caches.open(CACHE_NAME).then((cache) => cache.addAll(APP_SHELL)).catch(() => {})
  );
  self.skipWaiting();
});

self.addEventListener("activate", (event) => {
  event.waitUntil(
    caches.keys().then((keys) =>
      Promise.all(keys.filter((k) => k !== CACHE_NAME).map((k) => caches.delete(k)))
    )
  );
  self.clients.claim();
});

self.addEventListener("fetch", (event) => {
  const req = event.request;
  if (req.method !== "GET") return; // jamais les POST (uploads, ajouts de docs/événements...)
  const url = new URL(req.url);

  // Fichiers de l'appli (CSS/JS/icônes) : servis depuis le cache pour un chargement
  // instantané, et rafraîchis en arrière-plan dès que le réseau répond.
  if (url.origin === self.location.origin && APP_SHELL.includes(url.pathname)) {
    event.respondWith(
      caches.match(req).then((cached) => {
        const network = fetch(req)
          .then((res) => {
            caches.open(CACHE_NAME).then((cache) => cache.put(req, res.clone()));
            return res;
          })
          .catch(() => cached);
        return cached || network;
      })
    );
    return;
  }

  // Pages et données (matchs, documents, calendrier...) : toujours le réseau en
  // priorité, pour ne jamais afficher des stats périmées. Le cache ne sert que
  // de filet de sécurité si le téléphone perd la connexion.
  event.respondWith(fetch(req).catch(() => caches.match(req)));
});
""" % {"v": ASSET_VERSION}

# ---------------------------------------------------------------------------
# PWA — rend le site installable sur l'écran d'accueil (iPhone/Android), sans
# passer par l'App Store/Play Store. Manifest généré dynamiquement (pas un
# fichier static) pour reprendre le nom du club configuré par CLUB_NAME /
# CLUB_FULL_NAME. Les deux routes doivent rester accessibles sans connexion
# (voir PUBLIC_ENDPOINTS) : le navigateur les demande dès la page de login,
# avant que la personne ne soit authentifiée.
# ---------------------------------------------------------------------------
@app.route("/manifest.json")
def pwa_manifest():
    manifest = {
        "name": f"{CLUB_FULL_NAME} — Rugby Analytics",
        "short_name": CLUB_NAME or "Rugby Analytics",
        "description": "Analyse vidéo, documents, calendrier et suivi du staff, au même endroit.",
        "start_url": "/?source=pwa",
        "scope": "/",
        "display": "standalone",
        "orientation": "portrait-primary",
        "background_color": "#10181f",
        "theme_color": "#530d34",
        "icons": [
            {"src": url_for("static", filename="icon-192.png"), "sizes": "192x192", "type": "image/png", "purpose": "any"},
            {"src": url_for("static", filename="icon-512.png"), "sizes": "512x512", "type": "image/png", "purpose": "any"},
            {"src": url_for("static", filename="icon-maskable-512.png"), "sizes": "512x512", "type": "image/png", "purpose": "maskable"},
        ],
    }
    # mimetype précis (et pas juste application/json) : certains navigateurs sont
    # stricts sur ce point pour proposer l'installation de l'appli.
    return Response(json.dumps(manifest, ensure_ascii=False), mimetype="application/manifest+json")

@app.route("/sw.js")
def pwa_service_worker():
    resp = Response(SERVICE_WORKER_JS, mimetype="application/javascript")
    # Le fichier de service worker ne doit jamais rester en cache navigateur :
    # sinon une mise à jour du site ne serait jamais détectée par les téléphones
    # qui ont déjà installé l'appli.
    resp.headers["Cache-Control"] = "no-cache"
    return resp

@app.route("/demo/<token>")
def demo_login(token):
    if token != DEMO_TOKEN:
        flash("Lien de démonstration invalide.", "error")
        return redirect(url_for("login"))
    session.permanent = True
    session["logged_in"] = True
    session["is_admin"] = False
    session["demo_forced"] = True
    session["user_email"] = "démo"
    return redirect(url_for("landing", demo="1"))
@app.route("/login", methods=["GET", "POST"])
def login():
    next_url = request.values.get("next") or url_for("landing")
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        next_target = request.form.get("next") or url_for("landing")
        if email == ADMIN_EMAIL.lower() and password == ADMIN_PASSWORD:
            session.permanent = True
            session["logged_in"] = True
            session["is_admin"] = True
            session["is_player"] = False
            session["demo_forced"] = False
            session["user_email"] = email
            flash("Connecté.", "success")
            return redirect(next_target)
        for staff_email, staff_password in STAFF_ACCOUNTS:
            if email == staff_email and password == staff_password:
                session.permanent = True
                session["logged_in"] = True
                session["is_admin"] = False
                session["is_player"] = False
                session["demo_forced"] = False
                session["user_email"] = email
                flash("Connecté.", "success")
                return redirect(next_target)
        # Comptes joueurs : pas de mot de passe défini au départ. À la toute première
        # connexion, ce que le joueur tape dans le champ mot de passe DEVIENT son mot de
        # passe (pas d'email de confirmation possible sans serveur mail configuré) — voir
        # le message d'aide sur la page de connexion.
        db = get_db()
        player = db.execute("SELECT * FROM players WHERE email = %s", (email,)).fetchone()
        if player:
            if not password:
                flash("Merci d'indiquer un mot de passe.", "error")
                return render_template("login.html", next_url=next_url)
            if not player["password_hash"]:
                db.execute(
                    "UPDATE players SET password_hash = %s WHERE id = %s",
                    (generate_password_hash(password), player["id"]),
                )
                db.commit()
                session.permanent = True
                session["logged_in"] = True
                session["is_admin"] = False
                session["is_player"] = True
                session["player_id"] = player["id"]
                session["demo_forced"] = False
                session["user_email"] = email
                flash(f"Bienvenue {player['first_name']} ! Ton mot de passe vient d'être défini — ressers-t'en pour te reconnecter la prochaine fois.", "success")
                return redirect(url_for("player_home"))
            if check_password_hash(player["password_hash"], password):
                session.permanent = True
                session["logged_in"] = True
                session["is_admin"] = False
                session["is_player"] = True
                session["player_id"] = player["id"]
                session["demo_forced"] = False
                session["user_email"] = email
                flash("Connecté.", "success")
                return redirect(url_for("player_home") if next_target == url_for("landing") else next_target)
            flash("Email ou mot de passe incorrect.", "error")
            return render_template("login.html", next_url=next_url)
        flash("Email ou mot de passe incorrect.", "error")
    return render_template("login.html", next_url=next_url)
@app.route("/logout")
def logout():
    session.pop("logged_in", None)
    session.pop("is_admin", None)
    session.pop("is_player", None)
    session.pop("player_id", None)
    session.pop("demo_forced", None)
    session.pop("user_email", None)
    flash("Déconnecté.", "success")
    return redirect(url_for("login"))
class DB:
    """Petit adaptateur autour de psycopg2 pour garder l'écriture
    db.execute(sql, params).fetchall() / .fetchone() utilisée partout dans ce fichier,
    exactement comme avec sqlite3 avant. Les lignes se comportent comme des dictionnaires
    (row["colonne"]) grâce à RealDictCursor."""
    def __init__(self, conn):
        self._conn = conn
    def execute(self, query, params=()):
        cur = self._conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(query, params)
        return cur
    def commit(self):
        self._conn.commit()
    def close(self):
        self._conn.close()
def _connect():
    if not DATABASE_URL:
        raise RuntimeError(
            "DATABASE_URL n'est pas configurée. Ajoute-la dans les variables "
            "d'environnement Render (elle vient de ta base Neon)."
        )
    return psycopg2.connect(DATABASE_URL)
def get_db():
    if "db" not in g:
        g.db = DB(_connect())
    return g.db
@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()
def init_db():
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    conn = _connect()
    db = DB(conn)
    db.execute("""
        CREATE TABLE IF NOT EXISTS matches (
            id SERIAL PRIMARY KEY,
            created_at TEXT NOT NULL,
            match_date TEXT,
            own_team TEXT,
            opponent TEXT NOT NULL,
            competition TEXT,
            venue TEXT,
            own_team_tag TEXT,
            filename TEXT,
            total_instances INTEGER,
            stats_json TEXT,
            players_json TEXT,
            zones_json TEXT,
            instances_json TEXT
        )
    """)
    # Ajout de la colonne si la base existe déjà sans (anciennes versions du site)
    cur = db.execute(
        "SELECT column_name FROM information_schema.columns WHERE table_name = 'matches'"
    )
    cols = [r["column_name"] for r in cur.fetchall()]
    if "instances_json" not in cols:
        db.execute("ALTER TABLE matches ADD COLUMN instances_json TEXT")
    if "manual_stats_json" not in cols:
        db.execute("ALTER TABLE matches ADD COLUMN manual_stats_json TEXT")
    if "composition_json" not in cols:
        db.execute("ALTER TABLE matches ADD COLUMN composition_json TEXT")
    if "player_match_stats_json" not in cols:
        db.execute("ALTER TABLE matches ADD COLUMN player_match_stats_json TEXT")
    if "ubb_overview_json" not in cols:
        db.execute("ALTER TABLE matches ADD COLUMN ubb_overview_json TEXT")
    if "score_own" not in cols:
        db.execute("ALTER TABLE matches ADD COLUMN score_own INTEGER")
    if "score_opp" not in cols:
        db.execute("ALTER TABLE matches ADD COLUMN score_opp INTEGER")
    if "score_ht" not in cols:
        db.execute("ALTER TABLE matches ADD COLUMN score_ht TEXT")
    db.execute("""
        CREATE TABLE IF NOT EXISTS training_sessions (
            id SERIAL PRIMARY KEY,
            created_at TEXT NOT NULL,
            session_date TEXT,
            items_json TEXT
        )
    """)
    # Rapport hebdomadaire Pro D2 (fichier Excel) : une seule ligne à la fois, chaque
    # nouvel import remplace le précédent (pas d'historique semaine par semaine).
    db.execute("""
        CREATE TABLE IF NOT EXISTS prod2_reports (
            id SERIAL PRIMARY KEY,
            uploaded_at TEXT NOT NULL,
            filename TEXT,
            data_json TEXT
        )
    """)
    # Prochain adversaire sélectionné pour la page "Prochain match" : une seule ligne à la
    # fois (comme prod2_reports), remplacée à chaque changement de sélection.
    db.execute("""
        CREATE TABLE IF NOT EXISTS next_opponent (
            id SERIAL PRIMARY KEY,
            team TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    """)
    # Comptes joueurs : espace limité (planning en lecture seule, documents qui leur sont
    # partagés, leurs propres stats). Groupes (Avants/Trois-quarts...) génériques et gérables
    # par l'admin, utilisés à la fois pour classer les joueurs et pour cibler le partage
    # de documents.
    db.execute("""
        CREATE TABLE IF NOT EXISTS player_groups (
            id SERIAL PRIMARY KEY,
            name TEXT UNIQUE NOT NULL,
            created_at TEXT NOT NULL
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS players (
            id SERIAL PRIMARY KEY,
            first_name TEXT NOT NULL,
            last_name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT,
            group_id INTEGER REFERENCES player_groups(id) ON DELETE SET NULL,
            created_at TEXT NOT NULL
        )
    """)
    # Espace Documents du staff : dossiers libres + fichiers stockés DANS la base
    # PostgreSQL (colonne BYTEA) pour survivre aux redéploiements Render (le disque
    # du plan gratuit n'est pas persistant). Les vidéos lourdes passent par des
    # liens (kind='link') plutôt que des fichiers.
    db.execute("""
        CREATE TABLE IF NOT EXISTS doc_folders (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            parent_id INTEGER REFERENCES doc_folders(id) ON DELETE CASCADE,
            created_by TEXT,
            created_at TEXT NOT NULL
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS documents (
            id SERIAL PRIMARY KEY,
            folder_id INTEGER REFERENCES doc_folders(id) ON DELETE CASCADE,
            kind TEXT NOT NULL DEFAULT 'file',
            filename TEXT,
            mimetype TEXT,
            size_bytes INTEGER,
            data BYTEA,
            url TEXT,
            title TEXT,
            uploaded_by TEXT,
            uploaded_at TEXT NOT NULL
        )
    """)
    # Partage des documents avec les joueurs : ajouté après coup avec ALTER (la table
    # documents existe déjà en prod) plutôt que dans le CREATE TABLE ci-dessus.
    # visibility : 'staff' (comportement historique, réservé au staff) | 'players'
    # (tous les joueurs) | 'group' (un groupe précis) | 'player' (un joueur précis).
    db.execute("ALTER TABLE documents ADD COLUMN IF NOT EXISTS visibility TEXT NOT NULL DEFAULT 'staff'")
    db.execute("ALTER TABLE documents ADD COLUMN IF NOT EXISTS shared_group_id INTEGER REFERENCES player_groups(id) ON DELETE SET NULL")
    db.execute("ALTER TABLE documents ADD COLUMN IF NOT EXISTS shared_player_id INTEGER REFERENCES players(id) ON DELETE SET NULL")
    # Calendrier du staff : événements divers (réunions, déplacements, rendez-vous...),
    # distincts des matchs/entraînements qui restent gérés ailleurs sur le site.
    db.execute("""
        CREATE TABLE IF NOT EXISTS calendar_events (
            id SERIAL PRIMARY KEY,
            title TEXT NOT NULL,
            description TEXT,
            event_date TEXT NOT NULL,
            event_time TEXT,
            created_by TEXT,
            created_at TEXT NOT NULL
        )
    """)
    # Cahier des charges : suivi de tâches/exigences du staff, à 3 statuts.
    db.execute("""
        CREATE TABLE IF NOT EXISTS charges_items (
            id SERIAL PRIMARY KEY,
            title TEXT NOT NULL,
            description TEXT,
            status TEXT NOT NULL DEFAULT 'a_faire',
            created_by TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT
        )
    """)
    # Onglet par joueur dans le cahier des charges : NULL = tâche générale (comportement
    # historique) ; renseigné = suivi individuel propre à ce joueur (visible staff
    # uniquement, jamais montré au joueur — voir cahier_charges()).
    db.execute("ALTER TABLE charges_items ADD COLUMN IF NOT EXISTS player_id INTEGER REFERENCES players(id) ON DELETE CASCADE")
    # P.P.I.D (Projet Personnalisé Individuel du joueur) : reprise numérique du cahier
    # d'entraînement papier du club (mêmes 3 briques : profil par poste, évaluation rugby,
    # évaluation physique). ppid_position est plus fin que group_id (Avants/Trois-quarts) —
    # ex. "3LC" vs "3LA" — et sert uniquement à mettre en évidence la bonne colonne du
    # tableau de référence "Profil par poste" ; renseigné à la main par l'admin.
    db.execute("ALTER TABLE players ADD COLUMN IF NOT EXISTS ppid_position TEXT")
    # Une ligne = un point d'étape (3 à 4 par saison, comme sur le cahier papier) : notes
    # + commentaires par catégorie technique (JSON en texte plutôt qu'une colonne par
    # catégorie, pour ne pas avoir à migrer le schéma si les catégories évoluent un jour),
    # plus objectifs et entraînements spécifiques pour la période.
    db.execute("""
        CREATE TABLE IF NOT EXISTS ppid_rugby_evals (
            id SERIAL PRIMARY KEY,
            player_id INTEGER NOT NULL REFERENCES players(id) ON DELETE CASCADE,
            period_label TEXT NOT NULL,
            eval_date TEXT,
            ratings TEXT NOT NULL DEFAULT '{}',
            objectifs TEXT,
            entrainements TEXT,
            created_by TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT
        )
    """)
    # Même principe pour le physique, avec deux notes par catégorie et par période
    # (coach / auto-évaluation) stockées ensemble dans le même JSON : le staff renseigne
    # "coach", le joueur renseigne lui-même "auto" depuis son espace (voir
    # ppid_physical_auto_update()) — jamais l'inverse.
    db.execute("""
        CREATE TABLE IF NOT EXISTS ppid_physical_evals (
            id SERIAL PRIMARY KEY,
            player_id INTEGER NOT NULL REFERENCES players(id) ON DELETE CASCADE,
            period_label TEXT NOT NULL,
            eval_date TEXT,
            ratings TEXT NOT NULL DEFAULT '{}',
            commentaires TEXT,
            axe_musculation TEXT,
            axe_terrain TEXT,
            created_by TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT
        )
    """)
    # Cahier d'entretien : journal chronologique des briefings, retours/préparations de
    # match et entretiens individuels avec le joueur (page 2 du cahier papier du club).
    # Distinct du suivi de tâches (charges_items, staff uniquement) : ici, rédigé par le
    # staff mais VISIBLE par le joueur (voir player_evaluations()), comme le reste du PPID.
    db.execute("""
        CREATE TABLE IF NOT EXISTS ppid_entretiens (
            id SERIAL PRIMARY KEY,
            player_id INTEGER NOT NULL REFERENCES players(id) ON DELETE CASCADE,
            entretien_date TEXT NOT NULL,
            entretien_type TEXT NOT NULL,
            notes TEXT,
            created_by TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT
        )
    """)
    # Groupes de joueurs par défaut (l'admin peut en ajouter d'autres ensuite).
    for default_group in ("Avants", "Trois-quarts"):
        db.execute(
            "INSERT INTO player_groups (name, created_at) VALUES (%s, %s) ON CONFLICT (name) DO NOTHING",
            (default_group, datetime.utcnow().isoformat()),
        )
    db.commit()
    db.close()
@app.route("/")
def landing():
    return render_template("landing.html")
@app.route("/matchs")
def index():
    db = get_db()
    rows = db.execute("SELECT * FROM matches ORDER BY COALESCE(match_date, created_at) DESC, id DESC").fetchall()
    opponents = sorted({r["opponent"] for r in rows})
    competitions = sorted({r["competition"] for r in rows if r["competition"]})
    return render_template("index.html", matches=rows, opponents=opponents, competitions=competitions)
@app.route("/upload", methods=["GET", "POST"])
@admin_required
def upload():
    if request.method == "GET":
        return render_template("upload.html")
    file = request.files.get("xml_file")
    opponent = request.form.get("opponent", "").strip()
    if not file or file.filename == "":
        flash("Merci de sélectionner un fichier XML Sportscode.", "error")
        return redirect(url_for("upload"))
    if not opponent:
        flash("Merci d'indiquer le nom de l'adversaire.", "error")
        return redirect(url_for("upload"))
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    safe_name = f"{ts}_{file.filename}"
    save_path = os.path.join(UPLOAD_DIR, safe_name)
    file.save(save_path)
    try:
        parsed = parse_sportscode_xml(save_path)
    except Exception as exc:
        flash(f"Erreur lors de la lecture du fichier XML : {exc}", "error")
        return redirect(url_for("upload"))
    stats, players = aggregate_match_stats(parsed["instances"])
    zones = aggregate_zones(parsed["instances"])
    db = get_db()
    cur = db.execute(
        """INSERT INTO matches
           (created_at, match_date, own_team, opponent, competition, venue, own_team_tag,
            filename, total_instances, stats_json, players_json, zones_json, instances_json)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
           RETURNING id""",
        (
            datetime.utcnow().isoformat(),
            request.form.get("match_date") or None,
            request.form.get("own_team") or parsed["own_team_tag"],
            opponent,
            request.form.get("competition") or None,
            request.form.get("venue") or None,
            parsed["own_team_tag"],
            file.filename,
            len(parsed["instances"]),
            json.dumps(stats),
            json.dumps(players),
            json.dumps(zones),
            json.dumps(parsed["instances"]),
        ),
    )
    match_id = cur.fetchone()["id"]
    db.commit()
    flash("Match importé avec succès.", "success")
    return redirect(url_for("match_detail", match_id=match_id))
@app.route("/upload-ubb", methods=["GET", "POST"])
@admin_required
def upload_ubb():
    """Import d'un match UBB : un seul fichier XML "Équipe - Action" (convention
    différente de celle de Nice, voir parser_ubb.py). Le score se saisit à la main,
    il n'est pas reconstruit depuis le fichier."""
    if request.method == "GET":
        return render_template("upload_ubb.html")
    file = request.files.get("xml_file")
    opponent = request.form.get("opponent", "").strip()
    if not file or file.filename == "":
        flash("Merci de sélectionner un fichier XML Sportscode.", "error")
        return redirect(url_for("upload_ubb"))
    if not opponent:
        flash("Merci d'indiquer le nom exact de l'adversaire tel qu'il apparaît dans le fichier.", "error")
        return redirect(url_for("upload_ubb"))
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    safe_name = f"{ts}_{file.filename}"
    save_path = os.path.join(UPLOAD_DIR, safe_name)
    file.save(save_path)
    try:
        instances = parse_ubb_xml(save_path)
        overview = compute_ubb_overview(instances, own_team="Union Bordeaux Begles", opp_team=opponent)
    except Exception as exc:
        flash(f"Erreur lors de la lecture du fichier XML : {exc}", "error")
        return redirect(url_for("upload_ubb"))

    def _to_int(raw):
        raw = (raw or "").strip()
        if raw == "":
            return None
        try:
            return int(raw)
        except ValueError:
            return None

    db = get_db()
    cur = db.execute(
        """INSERT INTO matches
           (created_at, match_date, own_team, opponent, competition, venue, own_team_tag,
            filename, total_instances, ubb_overview_json, score_own, score_opp, score_ht)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
           RETURNING id""",
        (
            datetime.utcnow().isoformat(),
            request.form.get("match_date") or None,
            CLUB_FULL_NAME,
            opponent,
            request.form.get("competition") or None,
            request.form.get("venue") or None,
            "Union Bordeaux Begles",
            file.filename,
            len(instances),
            json.dumps(overview),
            _to_int(request.form.get("score_own")),
            _to_int(request.form.get("score_opp")),
            request.form.get("score_ht") or None,
        ),
    )
    match_id = cur.fetchone()["id"]
    db.commit()
    flash("Match UBB importé avec succès.", "success")
    return redirect(url_for("match_detail", match_id=match_id))
def _row_to_match(row):
    m = dict(row)
    m["stats"] = json.loads(m.pop("stats_json") or "{}")
    m["players"] = json.loads(m.pop("players_json") or "{}")
    m["zones"] = json.loads(m.pop("zones_json") or "{}")
    m["instances"] = json.loads(m.pop("instances_json") or "[]")
    m["manual_stats"] = json.loads(m.pop("manual_stats_json", None) or "{}")
    m["composition"] = json.loads(m.pop("composition_json", None) or "[]")
    m["player_match_stats"] = json.loads(m.pop("player_match_stats_json", None) or "{}")
    m["ubb_overview"] = json.loads(m.pop("ubb_overview_json", None) or "{}")
    return m
def _get_match_or_404(match_id):
    db = get_db()
    row = db.execute("SELECT * FROM matches WHERE id = %s", (match_id,)).fetchone()
    if row is None:
        abort(404)
    return _row_to_match(row)
def _no_instances_guard(match):
    """Les matchs importés avant la mise à jour 'secteurs' n'ont pas d'instances brutes stockées."""
    return not match["instances"]
@app.route("/match/<int:match_id>")
def match_detail(match_id):
    match = _get_match_or_404(match_id)
    if match.get("ubb_overview"):
        return render_template("match_ubb.html", match=match, ov=match["ubb_overview"])
    sections = []
    for section_name, cats in CATEGORY_SECTIONS.items():
        section_rows = []
        for cat in cats:
            data = match["stats"].get(cat)
            if not data:
                continue
            own = data.get("own", {"count": 0})
            adv = data.get("adverse", {"count": 0})
            neutral = data.get("neutral", {"count": 0})
            if own.get("count", 0) == 0 and adv.get("count", 0) == 0 and neutral.get("count", 0) == 0:
                continue
            help_text = CATEGORY_HELP.get(cat, "")
            section_rows.append({"category": cat, "own": own, "adverse": adv, "neutral": neutral, "help": help_text})
        if section_rows:
            sections.append({
                "name": section_name,
                "icon": SECTION_ICONS.get(section_name, ""),
                "help": SECTION_HELP.get(section_name, ""),
                "rows": section_rows,
            })
    top_players = sorted(match["players"].items(), key=lambda x: -x[1]["count"])[:15]
    highlights = generate_highlights(match["stats"], match["own_team"], match["opponent"])
    radar = compute_radar_metrics(match["stats"])
    score = None
    phase_timing = None
    dashboard = None
    baseline = None
    if match["instances"]:
        score = compute_score(match["instances"])
        phase_timing = compute_phase_timing(match["instances"])
        dashboard = compute_overview_dashboard(match["instances"], score)
        matches_with_instances, _, _, _ = _season_context()
        baseline = compute_match_baseline(matches_with_instances, exclude_id=match_id)
    return render_template(
        "match.html", match=match, sections=sections, top_players=top_players,
        highlights=highlights, radar=radar, score=score, phase_timing=phase_timing,
        phase_icons=PHASE_ICONS, phase_help=PHASE_HELP, dashboard=dashboard,
        baseline=baseline,
        has_instances=not _no_instances_guard(match),
    )
@app.route("/match/<int:match_id>/attaque")
def match_attaque(match_id):
    match = _get_match_or_404(match_id)
    if _no_instances_guard(match):
        flash("Ce match a été importé avant la mise à jour détaillée par secteur : réimporte le fichier XML pour voir cette page.", "error")
        return redirect(url_for("match_detail", match_id=match_id))
    attack = compute_attack_sector(match["instances"], "own")
    matches_with_instances, _, _, _ = _season_context()
    baselines = compute_sector_baselines(matches_with_instances, exclude_id=match_id)
    return render_template("match_attaque.html", match=match, data=attack, phase_icons=PHASE_ICONS,
                           baseline=(baselines or {}).get("attaque"))
@app.route("/match/<int:match_id>/defense")
def match_defense(match_id):
    match = _get_match_or_404(match_id)
    if _no_instances_guard(match):
        flash("Ce match a été importé avant la mise à jour détaillée par secteur : réimporte le fichier XML pour voir cette page.", "error")
        return redirect(url_for("match_detail", match_id=match_id))
    attack_adv = compute_attack_sector(match["instances"], "adverse")
    defense = compute_defense_sector(match["instances"], "adverse")
    matches_with_instances, _, _, _ = _season_context()
    baselines = compute_sector_baselines(matches_with_instances, exclude_id=match_id)
    return render_template("match_defense.html", match=match, data=attack_adv, defense=defense, phase_icons=PHASE_ICONS,
                           baseline=(baselines or {}).get("defense"))
@app.route("/match/<int:match_id>/ruck")
def match_ruck(match_id):
    match = _get_match_or_404(match_id)
    if _no_instances_guard(match):
        flash("Ce match a été importé avant la mise à jour détaillée par secteur : réimporte le fichier XML pour voir cette page.", "error")
        return redirect(url_for("match_detail", match_id=match_id))
    ruck = compute_ruck_sector(match["instances"])
    matches_with_instances, _, _, _ = _season_context()
    baselines = compute_sector_baselines(matches_with_instances, exclude_id=match_id)
    return render_template("match_ruck.html", match=match, data=ruck, phase_icons=PHASE_ICONS,
                           baseline=(baselines or {}).get("ruck"))
@app.route("/match/<int:match_id>/touches")
def match_touches(match_id):
    match = _get_match_or_404(match_id)
    if _no_instances_guard(match):
        flash("Ce match a été importé avant la mise à jour détaillée par secteur : réimporte le fichier XML pour voir cette page.", "error")
        return redirect(url_for("match_detail", match_id=match_id))
    lineout = compute_lineout_detail(match["instances"])
    matches_with_instances, _, _, _ = _season_context()
    baselines = compute_sector_baselines(matches_with_instances, exclude_id=match_id)
    return render_template("match_touches.html", match=match, data=lineout,
                           baseline=(baselines or {}).get("touches"))
@app.route("/match/<int:match_id>/melee")
def match_melee(match_id):
    match = _get_match_or_404(match_id)
    if _no_instances_guard(match):
        flash("Ce match a été importé avant la mise à jour détaillée par secteur : réimporte le fichier XML pour voir cette page.", "error")
        return redirect(url_for("match_detail", match_id=match_id))
    scrum = compute_scrum_detail(match["instances"])
    matches_with_instances, _, _, _ = _season_context()
    baselines = compute_sector_baselines(matches_with_instances, exclude_id=match_id)
    return render_template("match_melee.html", match=match, data=scrum, phase_icons=PHASE_ICONS,
                           baseline=(baselines or {}).get("melee"))
@app.route("/match/<int:match_id>/jap")
def match_jap(match_id):
    match = _get_match_or_404(match_id)
    if _no_instances_guard(match):
        flash("Ce match a été importé avant la mise à jour détaillée par secteur : réimporte le fichier XML pour voir cette page.", "error")
        return redirect(url_for("match_detail", match_id=match_id))
    kicking = compute_kicking_detail(match["instances"])
    matches_with_instances, _, _, _ = _season_context()
    baselines = compute_sector_baselines(matches_with_instances, exclude_id=match_id)
    return render_template("match_jap.html", match=match, data=kicking,
                           baseline=(baselines or {}).get("jap"))
@app.route("/match/<int:match_id>/jap/manual", methods=["POST"])
@admin_required
def match_jap_manual(match_id):
    match = _get_match_or_404(match_id)
    manual = match.get("manual_stats") or {}
    def _to_int(raw):
        raw = (raw or "").strip()
        if raw == "":
            return None
        try:
            return int(raw)
        except ValueError:
            return None
    manual["kick_distance_own"] = _to_int(request.form.get("kick_distance_own"))
    manual["kick_distance_adverse"] = _to_int(request.form.get("kick_distance_adverse"))
    db = get_db()
    db.execute("UPDATE matches SET manual_stats_json = %s WHERE id = %s", (json.dumps(manual), match_id))
    db.commit()
    flash("Données jeu au pied mises à jour.", "success")
    return redirect(url_for("match_jap", match_id=match_id))
@app.route("/match/<int:match_id>/joueurs")
def match_joueurs(match_id):
    match = _get_match_or_404(match_id)
    if _no_instances_guard(match):
        flash("Ce match a été importé avant la mise à jour détaillée par secteur : réimporte le fichier XML pour voir cette page.", "error")
        return redirect(url_for("match_detail", match_id=match_id))
    attack_table = compute_player_attack_table(match["instances"])
    defense_table = compute_player_defense_table(match["instances"])
    ruck_table = compute_player_ruck_table(match["instances"])
    matches_with_instances, _, _, _ = _season_context()
    player_baselines = compute_player_season_baselines(matches_with_instances, exclude_id=match_id)
    player_cards = attach_overview_highlights(build_player_cards(
        attack_table, defense_table, ruck_table, composition=match.get("composition")))
    return render_template("match_joueurs.html", match=match, attack_table=attack_table,
                           defense_table=defense_table, ruck_table=ruck_table,
                           player_baselines=player_baselines, player_cards=player_cards)
@app.route("/match/<int:match_id>/composition", methods=["GET", "POST"])
def match_composition(match_id):
    match = _get_match_or_404(match_id)
    if request.method == "POST":
        if not session.get("is_admin"):
            flash("Connecte-toi pour valider la composition.", "error")
            return redirect(url_for("login", next=request.path))
        selected = request.form.getlist("player")
        selected = (selected + [""] * 23)[:23]  # toujours 23 emplacements (n°1 à n°23)
        minutes_in = (request.form.getlist("minutes") + [""] * 23)[:23]
        yellow_in = (request.form.getlist("yellow") + [""] * 23)[:23]
        red_in = (request.form.getlist("red") + [""] * 23)[:23]
        def _to_int(v):
            try:
                return max(0, int(v))
            except (TypeError, ValueError):
                return 0
        player_stats = {}
        for i, name in enumerate(selected):
            if not name:
                continue
            player_stats[name] = {
                "minutes": _to_int(minutes_in[i]),
                "yellow": _to_int(yellow_in[i]),
                "red": _to_int(red_in[i]),
            }
        db = get_db()
        db.execute(
            "UPDATE matches SET composition_json = %s, player_match_stats_json = %s WHERE id = %s",
            (json.dumps(selected), json.dumps(player_stats), match_id),
        )
        db.commit()
        flash("Composition enregistrée.", "success")
        return redirect(url_for("match_composition", match_id=match_id))
    slots = (match.get("composition") or [])[:23]
    slots = slots + [""] * (23 - len(slots))
    all_players = [
        {"position": position, "players": [{"name": p, "jiff": is_jiff(p)} for p in SQUAD_ROSTER[position]]}
        for position in SQUAD_POSITION_ORDER
    ]
    filled_count = sum(1 for p in slots if p)
    player_stats = match.get("player_match_stats") or {}
    return render_template("match_composition.html", match=match, all_players=all_players,
                           slots=slots, filled_count=filled_count, player_stats=player_stats)
@app.route("/match/<int:match_id>/transition")
def match_transition(match_id):
    match = _get_match_or_404(match_id)
    if _no_instances_guard(match):
        flash("Ce match a été importé avant la mise à jour détaillée par secteur : réimporte le fichier XML pour voir cette page.", "error")
        return redirect(url_for("match_detail", match_id=match_id))
    transition = compute_transition_sector(match["instances"])
    return render_template("match_transition.html", match=match, data=transition)
@app.route("/admin/export")
@admin_required
def export_data():
    """Télécharge une sauvegarde complète (JSON) de tous les matchs enregistrés.
    À utiliser avant toute mise à jour du site pour ne jamais perdre de données."""
    db = get_db()
    rows = db.execute("SELECT * FROM matches ORDER BY id").fetchall()
    data = [dict(r) for r in rows]
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    return Response(
        payload,
        mimetype="application/json",
        headers={"Content-Disposition": f"attachment; filename=rugby_analytics_sauvegarde_{ts}.json"},
    )
@app.route("/admin/import", methods=["GET", "POST"])
@admin_required
def import_data():
    """Recharge des matchs depuis un fichier de sauvegarde JSON (généré par 'Sauvegarder
    les données'). Sert à restaurer les données juste après un changement de base de
    données, ou si besoin de récupérer une ancienne sauvegarde."""
    if request.method == "GET":
        return render_template("admin_import.html")
    file = request.files.get("backup_file")
    if not file or file.filename == "":
        flash("Merci de sélectionner un fichier de sauvegarde JSON.", "error")
        return redirect(url_for("import_data"))
    try:
        data = json.loads(file.read().decode("utf-8"))
    except Exception as exc:
        flash(f"Fichier de sauvegarde invalide : {exc}", "error")
        return redirect(url_for("import_data"))
    db = get_db()
    imported = 0
    for m in data:
        db.execute(
            """INSERT INTO matches
               (created_at, match_date, own_team, opponent, competition, venue, own_team_tag,
                filename, total_instances, stats_json, players_json, zones_json, instances_json,
                manual_stats_json, composition_json)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (
                m.get("created_at"), m.get("match_date"), m.get("own_team"), m.get("opponent"),
                m.get("competition"), m.get("venue"), m.get("own_team_tag"), m.get("filename"),
                m.get("total_instances"), m.get("stats_json"), m.get("players_json"),
                m.get("zones_json"), m.get("instances_json"), m.get("manual_stats_json"),
                m.get("composition_json"),
            ),
        )
        imported += 1
    db.commit()
    flash(f"{imported} match(s) importé(s) depuis la sauvegarde.", "success")
    return redirect(url_for("index"))
@app.route("/match/<int:match_id>/delete", methods=["POST"])
@admin_required
def delete_match(match_id):
    db = get_db()
    db.execute("DELETE FROM matches WHERE id = %s", (match_id,))
    db.commit()
    flash("Match supprimé.", "success")
    return redirect(url_for("index"))
SECTOR_PAGE_META = {
    "attaque": {"title": "Attaque", "icon": "⚔️"},
    "defense": {"title": "Défense", "icon": "🛡️"},
    "discipline": {"title": "Discipline", "icon": "🟨"},
    "touches": {"title": "Touches", "icon": "🤾"},
    "melee": {"title": "Mêlée", "icon": "🔒"},
    "rucks": {"title": "Rucks", "icon": "🤝"},
    "jap": {"title": "Jeu au pied", "icon": "🦶"},
}
def _extract_journee(filename):
    """Numéro de journée extrait du nom du fichier Pro D2 (ex: '..._Journée 1_Finale.xlsx'
    -> 1), pour afficher la fraîcheur des données sans avoir à le ressaisir à la main."""
    if not filename:
        return None
    match = re.search(r"[Jj]ourn[ée]e\s*(\d+)", filename)
    return int(match.group(1)) if match else None
def _load_latest_prod2_row():
    db = get_db()
    return db.execute(
        "SELECT uploaded_at, filename, data_json FROM prod2_reports ORDER BY id DESC LIMIT 1"
    ).fetchone()
def _load_prod2_report():
    """Rapport Pro D2 le plus récent. Chaque import est conservé en base (voir
    _load_prod2_history) pour pouvoir suivre l'évolution d'une équipe semaine après
    semaine, mais toutes les pages "état actuel" du site n'affichent toujours que celui-ci.
    Renvoie None si aucun rapport n'a encore été importé."""
    row = _load_latest_prod2_row()
    if not row:
        return None
    return json.loads(row["data_json"])
def _load_prod2_meta():
    """Date d'import + numéro de journée du dernier rapport, pour afficher la fraîcheur des
    données sur les pages Adversaires (ex: 'Journée 3, importé le 14/07/2026 à 11h56')."""
    row = _load_latest_prod2_row()
    if not row:
        return None
    return {
        "uploaded_at": row["uploaded_at"],
        "filename": row["filename"],
        "journee": _extract_journee(row["filename"]),
    }
def _report_label(report_meta):
    if not report_meta:
        return "aucun rapport importé"
    parts = []
    if report_meta.get("journee"):
        parts.append(f"Journée {report_meta['journee']}")
    if report_meta.get("uploaded_at"):
        try:
            dt = datetime.fromisoformat(report_meta["uploaded_at"])
            parts.append(f"importé le {dt.strftime('%d/%m/%Y à %Hh%M')}")
        except ValueError:
            pass
    return ", ".join(parts) if parts else "dernier rapport importé"
def _load_prod2_history():
    """Tous les rapports Pro D2 importés depuis l'activation de l'historique, du plus
    ancien au plus récent, sous la forme attendue par compute_team_position_history."""
    db = get_db()
    rows = db.execute(
        "SELECT uploaded_at, filename, data_json FROM prod2_reports ORDER BY uploaded_at ASC"
    ).fetchall()
    history = []
    for r in rows:
        journee = _extract_journee(r["filename"])
        label = f"J{journee}" if journee else (r["uploaded_at"] or "")[:10]
        history.append({"label": label, "data": json.loads(r["data_json"])})
    return history
# Comparaison "Nissa vs adversaire" sur la fiche Vue d'ensemble : uniquement des pourcentages
# (jamais des totaux bruts), pour ne jamais mélanger un cumul sur nos matchs codés avec un
# cumul sur toute la saison Pro D2 de l'adversaire — deux échelles différentes qui rendraient
# une comparaison de totaux absurde. Même en pourcentage, nos stats viennent de notre propre
# codage vidéo (Sportscode) et celles de l'adversaire du prestataire officiel Pro D2 : les
# méthodologies de calcul peuvent différer légèrement, d'où l'avertissement affiché avec.
NISSA_VS_OPPONENT_METRICS = [
    ("Possession (%)", "possession", "Possession", "% Possession"),
    ("Plaquages réussis (%)", "tackle_pct", "Défense", "Plaquages réussis %"),
    ("Touches gagnées (%)", "lineout_pct", "Touches", "Touches gagnées %"),
    ("Mêlées gagnées (%)", "scrum_pct", "Mêlées", "Mêlées gagnées %"),
    ("Duels aériens gagnés (%)", "duels_aeriens_pct", "Duels aériens", "Duels aériens gagnés %"),
]
def _compute_nissa_vs_opponent(profile):
    """Compare nos propres stats de saison (calculées sur tous nos matchs codés) à celles de
    l'adversaire dans le rapport Pro D2, limité aux indicateurs en pourcentage pour rester
    comparable malgré les échelles différentes. Renvoie None si on n'a pas encore de match
    codé cette saison."""
    matches_with_instances, selected, selected_ids, qs = _season_context()
    if not selected:
        return None
    instances = _season_instances(selected)
    our_kpis = compute_match_kpis(instances)
    poss_own = poss_adv = 0
    for m in selected:
        poss = m["stats"].get("Possession", {})
        poss_own += poss.get("own", {}).get("duration", 0)
        poss_adv += poss.get("adverse", {}).get("duration", 0)
    poss_total = poss_own + poss_adv
    our_kpis["possession"] = round(poss_own / poss_total * 100, 1) if poss_total else None
    rows = []
    for label, our_key, sheet, column in NISSA_VS_OPPONENT_METRICS:
        our_value = our_kpis.get(our_key)
        their_value = profile["categories"].get(sheet, {}).get(column)
        if our_value is None and their_value is None:
            continue
        rows.append({"label": label, "own": our_value, "opponent": their_value})
    return rows
def _team_profile_or_404(team, report=None):
    if report is None:
        report = _load_prod2_report()
    if not report:
        abort(404)
    profile = get_team_profile(report, team)
    if profile["classement"] is None:
        abort(404)
    return profile
def _sector_sheets(profile, sector_key):
    sheets = []
    for sheet_name in SECTOR_SHEETS[sector_key]:
        team_row = profile["categories"].get(sheet_name, {})
        avg_row = profile["league_avg"].get(sheet_name, {})
        sheets.append({
            "name": sheet_name,
            "team_row": team_row,
            "avg_row": avg_row,
            "rows": build_compare_rows(sheet_name, team_row, avg_row),
        })
    return sheets
@app.route("/opponents")
def opponents():
    report = _load_prod2_report()
    report_meta = _load_prod2_meta()
    classement = get_classement_table(report) if report else []
    return render_template(
        "opponents.html", classement=classement, has_report=report is not None,
        report_label=_report_label(report_meta),
    )
@app.route("/opponents/<team>")
def opponent_detail(team):
    report = _load_prod2_report()
    if not report:
        abort(404)
    profile = _team_profile_or_404(team, report=report)
    possession = profile["categories"].get("Possession", {})
    possession_avg = profile["league_avg"].get("Possession", {})
    possession_rows = build_compare_rows("Possession", possession, possession_avg)
    team_trends = compute_team_trends(report).get(team, {"strengths": [], "weaknesses": []})
    kpi_profile = compute_team_kpi_profile(report, team)
    report_meta = _load_prod2_meta()
    history = compute_team_position_history(_load_prod2_history(), team)
    nissa_compare = _compute_nissa_vs_opponent(profile)
    return render_template(
        "opponent.html", team=team, profile=profile,
        possession=possession, possession_avg=possession_avg, possession_rows=possession_rows,
        team_trends=team_trends, kpi_profile=kpi_profile, report_label=_report_label(report_meta),
        history=history, nissa_compare=nissa_compare,
        team_mode=True, active="overview",
    )
@app.route("/opponents/<team>/joueurs")
def opponent_joueurs(team):
    profile = _team_profile_or_404(team)
    groups = compute_player_groups(profile["players"])
    most_used = compute_most_used_players(profile["players"])
    threats = compute_team_threats(profile["players"])
    report_meta = _load_prod2_meta()
    return render_template(
        "opponent_joueurs.html", team=team, profile=profile, groups=groups, most_used=most_used,
        threats=threats, report_label=_report_label(report_meta),
        team_mode=True, active="joueurs",
    )
@app.route("/opponents/tendances")
def opponents_trends():
    report = _load_prod2_report()
    if not report:
        flash("Importe d'abord le rapport Pro D2 pour voir les tendances.", "error")
        return redirect(url_for("opponents"))
    trends = compute_team_trends(report)
    report_meta = _load_prod2_meta()
    return render_template(
        "opponent_trends.html", teams=report["team_names"], trends=trends,
        report_label=_report_label(report_meta),
    )
@app.route("/opponents/<team>/<sector>")
def opponent_sector(team, sector):
    if sector not in SECTOR_SHEETS:
        abort(404)
    profile = _team_profile_or_404(team)
    sheets = _sector_sheets(profile, sector)
    meta = SECTOR_PAGE_META[sector]
    report_meta = _load_prod2_meta()
    return render_template(
        "opponent_sector.html", team=team, profile=profile, sheets=sheets,
        team_mode=True, active=sector, page_title=meta["title"], page_icon=meta["icon"],
        report_label=_report_label(report_meta),
    )
def _head_to_head(matches_with_instances, team):
    """Historique de nos matchs déjà codés face à cet adversaire précis (comparaison de
    noms insensible à la casse), du plus ancien au plus récent, pour la page Prochain match."""
    team_norm = (team or "").strip().lower()
    rows = []
    wins = draws = losses = 0
    for m in matches_with_instances:
        if (m.get("opponent") or "").strip().lower() != team_norm:
            continue
        sc = compute_score(m["instances"])
        if sc["own"] > sc["adverse"]:
            result = "V"
            wins += 1
        elif sc["own"] < sc["adverse"]:
            result = "D"
            losses += 1
        else:
            result = "N"
            draws += 1
        rows.append({
            "match_id": m["id"], "date": m.get("match_date"), "competition": m.get("competition"),
            "own_score": sc["own"], "adverse_score": sc["adverse"], "result": result,
        })
    return {"rows": rows, "wins": wins, "draws": draws, "losses": losses, "total": len(rows)}
def _load_next_opponent():
    db = get_db()
    row = db.execute("SELECT team FROM next_opponent ORDER BY id DESC LIMIT 1").fetchone()
    return row["team"] if row else None
@app.route("/prochain-match")
def next_match():
    """Page de préparation de la semaine : notre forme récente (5 derniers matchs codés) +
    scouting du prochain adversaire (repris des pages Adversaires) + historique face à cette
    équipe, réunis au même endroit plutôt que de naviguer entre Bilan de saison et
    Adversaires séparément avant une réunion de préparation."""
    report = _load_prod2_report()
    report_meta = _load_prod2_meta()
    team_names = report["team_names"] if report else []
    selected_team = _load_next_opponent()
    profile = None
    possession_rows = None
    team_trends = None
    kpi_profile = None
    threats = None
    stale_team = False
    if report and selected_team:
        profile = get_team_profile(report, selected_team)
        if profile["classement"] is None:
            profile = None
            stale_team = True
        else:
            possession = profile["categories"].get("Possession", {})
            possession_avg = profile["league_avg"].get("Possession", {})
            possession_rows = build_compare_rows("Possession", possession, possession_avg)
            team_trends = compute_team_trends(report).get(selected_team, {"strengths": [], "weaknesses": []})
            kpi_profile = compute_team_kpi_profile(report, selected_team)
            threats = compute_team_threats(profile["players"])
    matches_with_instances, _, _, _ = _season_context()
    recent = matches_with_instances[-5:]
    recent_dashboard = compute_season_dashboard(recent) if recent else None
    recent_stats = aggregate_match_stats(_season_instances(recent))[0] if recent else {}
    head_to_head = _head_to_head(matches_with_instances, selected_team) if selected_team else None
    return render_template(
        "next_match.html", team_names=team_names, selected_team=selected_team,
        has_report=report is not None, stale_team=stale_team,
        profile=profile, possession_rows=possession_rows, team_trends=team_trends,
        kpi_profile=kpi_profile, threats=threats, report_label=_report_label(report_meta),
        recent_dashboard=recent_dashboard, recent_stats=recent_stats, recent_count=len(recent),
        head_to_head=head_to_head,
    )
@app.route("/prochain-match/set", methods=["POST"])
@admin_required
def next_match_set():
    team = (request.form.get("team") or "").strip()
    db = get_db()
    db.execute("DELETE FROM next_opponent")
    if team:
        db.execute(
            "INSERT INTO next_opponent (team, updated_at) VALUES (%s, %s)",
            (team, datetime.utcnow().isoformat()),
        )
    db.commit()
    flash("Prochain adversaire mis à jour." if team else "Prochain adversaire effacé.", "success")
    return redirect(url_for("next_match"))
@app.route("/admin/prod2/import", methods=["GET", "POST"])
@admin_required
def prod2_import():
    """Import du rapport hebdomadaire Pro D2 (fichier Excel). Le rapport le plus récent est
    toujours celui affiché partout sur le site, mais chaque import est conservé en base (au
    lieu d'écraser le précédent) pour pouvoir suivre l'évolution d'une équipe au fil des
    semaines (voir compute_team_position_history)."""
    if request.method == "GET":
        return render_template("admin_prod2_import.html")
    file = request.files.get("xlsx_file")
    if not file or file.filename == "":
        flash("Merci de sélectionner un fichier Excel Pro D2.", "error")
        return redirect(url_for("prod2_import"))
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    safe_name = f"{ts}_{file.filename}"
    save_path = os.path.join(UPLOAD_DIR, safe_name)
    file.save(save_path)
    try:
        report = parse_prod2_report(save_path)
    except Exception as exc:
        flash(f"Erreur lors de la lecture du fichier Excel : {exc}", "error")
        return redirect(url_for("prod2_import"))
    if not report["team_names"]:
        flash("Aucune équipe trouvée dans ce fichier — vérifie qu'il s'agit bien du bon export Pro D2.", "error")
        return redirect(url_for("prod2_import"))
    db = get_db()
    db.execute(
        "INSERT INTO prod2_reports (uploaded_at, filename, data_json) VALUES (%s, %s, %s)",
        (datetime.utcnow().isoformat(), file.filename, json.dumps(report)),
    )
    db.commit()
    flash(f"Rapport Pro D2 importé : {len(report['team_names'])} équipes.", "success")
    return redirect(url_for("opponents"))
def _season_context():
    """Matchs sélectionnés pour le cumul saison (case à cocher sur /season, conservée via
    ?m=id&m=id... sur toutes les pages secteur saison) + la query string à réutiliser dans
    les liens du sous-menu pour ne pas perdre la sélection en changeant d'onglet."""
    db = get_db()
    rows = db.execute(
        "SELECT * FROM matches ORDER BY COALESCE(match_date, created_at)"
    ).fetchall()
    all_matches = [_row_to_match(r) for r in rows]
    matches_with_instances = [m for m in all_matches if m["instances"]]
    selected_ids_param = request.args.getlist("m")
    if selected_ids_param:
        selected_ids = {int(x) for x in selected_ids_param if x.isdigit()}
        selected = [m for m in matches_with_instances if m["id"] in selected_ids]
    else:
        selected_ids = {m["id"] for m in matches_with_instances}
        selected = matches_with_instances
    qs = "&".join(f"m={i}" for i in sorted(selected_ids))
    return matches_with_instances, selected, selected_ids, qs
def _season_instances(selected):
    combined = []
    for m in selected:
        combined.extend(m["instances"])
    return combined
def _sum_manual(selected, key):
    vals = [m["manual_stats"].get(key) for m in selected if m.get("manual_stats", {}).get(key) is not None]
    return sum(vals) if vals else None
@app.route("/season")
def season():
    matches_with_instances, selected, selected_ids, qs = _season_context()
    dashboard = compute_season_dashboard(selected) if selected else None
    stats = aggregate_match_stats([i for m in selected for i in m["instances"]])[0] if selected else {}
    return render_template(
        "season.html",
        all_matches=matches_with_instances,
        selected_ids=selected_ids,
        selected_count=len(selected),
        total_count=len(matches_with_instances),
        dashboard=dashboard,
        stats=stats,
        season_mode=True, active="overview", qs=qs,
    )
@app.route("/season/attaque")
def season_attaque():
    _, selected, _, qs = _season_context()
    instances = _season_instances(selected)
    attack = compute_attack_sector(instances, "own") if instances else None
    if attack:
        attack["try_timing"] = compute_event_timing_multi(selected, "Essai", "own")
        attack["break_timing"] = compute_event_timing_multi(selected, "Break", "own")
    back3_trend = compute_back3_trend(selected) if selected else []
    return render_template("season_attaque.html", data=attack, back3_trend=back3_trend, season_mode=True,
                           active="attaque", qs=qs, selected_count=len(selected))
@app.route("/season/defense")
def season_defense():
    _, selected, _, qs = _season_context()
    instances = _season_instances(selected)
    attack_adv = compute_attack_sector(instances, "adverse") if instances else None
    if attack_adv:
        attack_adv["try_timing"] = compute_event_timing_multi(selected, "Essai", "adverse")
        attack_adv["break_timing"] = compute_event_timing_multi(selected, "Break", "adverse")
    defense = compute_defense_sector(instances, "adverse") if instances else None
    return render_template("season_defense.html", data=attack_adv, defense=defense, season_mode=True,
                           active="defense", qs=qs, selected_count=len(selected))
@app.route("/season/ruck")
def season_ruck():
    _, selected, _, qs = _season_context()
    instances = _season_instances(selected)
    ruck = compute_ruck_sector(instances) if instances else None
    return render_template("season_ruck.html", data=ruck, season_mode=True, active="ruck", qs=qs,
                           phase_icons=PHASE_ICONS, selected_count=len(selected))
@app.route("/season/touches")
def season_touches():
    _, selected, _, qs = _season_context()
    instances = _season_instances(selected)
    lineout = compute_lineout_detail(instances) if instances else None
    return render_template("season_touches.html", data=lineout, season_mode=True, active="touches", qs=qs,
                           selected_count=len(selected))
@app.route("/season/melee")
def season_melee():
    _, selected, _, qs = _season_context()
    instances = _season_instances(selected)
    scrum = compute_scrum_detail(instances) if instances else None
    return render_template("season_melee.html", data=scrum, season_mode=True, active="melee", qs=qs,
                           phase_icons=PHASE_ICONS, selected_count=len(selected))
@app.route("/season/jap")
def season_jap():
    _, selected, _, qs = _season_context()
    instances = _season_instances(selected)
    kicking = compute_kicking_detail(instances) if instances else None
    manual_totals = {
        "kick_distance_own": _sum_manual(selected, "kick_distance_own"),
        "kick_distance_adverse": _sum_manual(selected, "kick_distance_adverse"),
    }
    return render_template("season_jap.html", data=kicking, manual_totals=manual_totals, season_mode=True,
                           active="jap", qs=qs, selected_count=len(selected))
@app.route("/season/joueurs")
def season_joueurs():
    _, selected, _, qs = _season_context()
    instances = _season_instances(selected)
    groups = compute_squad_season_stats(instances, selected)
    attack_table = compute_player_attack_table(instances)
    defense_table = compute_player_defense_table(instances)
    ruck_table = compute_player_ruck_table(instances)
    player_cards = attach_overview_highlights(build_player_cards(attack_table, defense_table, ruck_table))
    return render_template("season_joueurs.html", groups=groups, season_mode=True, active="joueurs",
                           qs=qs, selected_count=len(selected),
                           attack_table=attack_table, defense_table=defense_table,
                           ruck_table=ruck_table, player_cards=player_cards)
@app.route("/season/comparateur")
def season_comparateur():
    _, selected, selected_ids, qs = _season_context()
    instances = _season_instances(selected)
    player_a = request.args.get("a") or ""
    player_b = request.args.get("b") or ""
    comparison = None
    radar_svg = None
    if instances and player_a and player_b:
        comparison = compute_player_comparison(instances, player_a, player_b)
        radar_svg = compute_player_radar_svg(comparison["attack_rows"], comparison["defense_rows"],
                                              comparison["ruck_rows"])
    all_players = [
        {"position": position, "players": SQUAD_ROSTER[position]}
        for position in SQUAD_POSITION_ORDER
    ]
    return render_template(
        "season_comparateur.html", data=comparison, radar_svg=radar_svg, all_players=all_players,
        player_a=player_a, player_b=player_b, selected_ids=selected_ids,
        season_mode=True, active="comparateur", qs=qs, selected_count=len(selected),
    )
@app.route("/season/jiff")
def season_jiff():
    _, selected, _, qs = _season_context()
    jiff_data = compute_jiff_chart(selected)
    return render_template("season_jiff.html", data=jiff_data, season_mode=True, active="jiff", qs=qs,
                           selected_count=len(selected))
@app.route("/season/transition")
def season_transition():
    _, selected, _, qs = _season_context()
    instances = _season_instances(selected)
    transition = compute_transition_sector(instances) if instances else None
    return render_template("season_transition.html", data=transition, season_mode=True, active="transition", qs=qs,
                           selected_count=len(selected))
def _load_training_sessions():
    db = get_db()
    rows = db.execute(
        "SELECT * FROM training_sessions ORDER BY session_date DESC, id DESC"
    ).fetchall()
    sessions = []
    for r in rows:
        sessions.append({
            "id": r["id"],
            "session_date": r["session_date"],
            "items": json.loads(r["items_json"] or "[]"),
        })
    return sessions
@app.route("/season/entrainement", methods=["GET", "POST"])
def season_entrainement():
    if request.method == "POST":
        if not session.get("is_admin"):
            flash("Connecte-toi pour enregistrer une séance.", "error")
            return redirect(url_for("login", next=request.path))
        session_date = request.form.get("session_date", "").strip()
        if not session_date:
            flash("Merci d'indiquer une date de séance.", "error")
            return redirect(url_for("season_entrainement"))
        items = []
        for cat_name, subcats in TRAINING_TAXONOMY.items():
            for subcat_name, elements in subcats.items():
                for element in elements:
                    check_field = f"item::{cat_name}::{subcat_name}::{element}"
                    if not request.form.get(check_field):
                        continue
                    minutes_raw = request.form.get(f"minutes::{cat_name}::{subcat_name}::{element}", "").strip()
                    minutes = int(minutes_raw) if minutes_raw.isdigit() else None
                    items.append({
                        "category": cat_name, "subcategory": subcat_name,
                        "element": element, "minutes": minutes,
                    })
        if not items:
            flash("Sélectionne au moins un élément travaillé lors de la séance.", "error")
            return redirect(url_for("season_entrainement"))
        db = get_db()
        db.execute(
            "INSERT INTO training_sessions (created_at, session_date, items_json) VALUES (%s, %s, %s)",
            (datetime.utcnow().isoformat(), session_date, json.dumps(items)),
        )
        db.commit()
        flash("Séance enregistrée.", "success")
        return redirect(url_for("season_entrainement"))
    sessions = _load_training_sessions()
    weeks = group_training_sessions_by_period(sessions, "week")
    months = group_training_sessions_by_period(sessions, "month")
    return render_template(
        "season_entrainement.html", weeks=weeks, months=months, sessions=sessions, taxonomy=TRAINING_TAXONOMY,
        today=datetime.utcnow().strftime("%Y-%m-%d"),
        season_mode=True, active="entrainement", qs="",
    )
@app.route("/season/entrainement/<int:session_id>/delete", methods=["POST"])
@admin_required
def delete_training_session(session_id):
    db = get_db()
    db.execute("DELETE FROM training_sessions WHERE id = %s", (session_id,))
    db.commit()
    flash("Séance supprimée.", "success")
    return redirect(url_for("season_entrainement"))
@app.route("/season/analyse")
def season_analyse():
    _, selected, _, qs = _season_context()
    analysis = compute_win_loss_analysis(selected)
    return render_template(
        "season_analyse.html", analysis=analysis, season_mode=True, active="analyse",
        qs=qs, selected_count=len(selected),
    )
# ---------------------------------------------------------------------------
# ESPACE DOCUMENTS — plateforme centrale du staff
# ---------------------------------------------------------------------------
# Chaque membre du staff (connecté avec son compte) peut créer des dossiers,
# déposer des fichiers (PDF, présentations, images, Excel... max 30 Mo) et
# ajouter des liens vidéo (Hudl, YouTube, Drive...). Les fichiers sont stockés
# dans PostgreSQL : ils survivent aux redéploiements Render.
# Règles : tout le staff peut déposer ; chacun peut supprimer SES dépôts ;
# l'admin peut tout supprimer ; le mode démo ne voit rien (voir gate_optional_features).

DOC_TYPE_ICONS = {
    "pdf": "📕", "doc": "📘", "docx": "📘", "ppt": "📙", "pptx": "📙", "key": "📙",
    "xls": "📗", "xlsx": "📗", "csv": "📗", "png": "🖼️", "jpg": "🖼️", "jpeg": "🖼️",
    "gif": "🖼️", "heic": "🖼️", "webp": "🖼️", "mp4": "🎬", "mov": "🎬", "zip": "🗜️",
    "txt": "📄", "xml": "📄",
}
# Extensions dont l'aperçu peut s'ouvrir directement dans le navigateur.
DOC_INLINE_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "gif", "webp", "txt"}

def _doc_ext(filename):
    return (filename or "").rsplit(".", 1)[-1].lower() if "." in (filename or "") else ""

def _doc_icon(doc):
    if doc["kind"] == "link":
        return "🔗"
    return DOC_TYPE_ICONS.get(_doc_ext(doc["filename"]), "📄")

def _doc_can_delete(row):
    """L'admin supprime tout ; un membre du staff supprime ce qu'il a déposé lui-même."""
    if session.get("is_admin"):
        return True
    email = session.get("user_email", "")
    return bool(email) and (row.get("uploaded_by") or row.get("created_by")) == email

def _folder_or_404(db, folder_id):
    folder = db.execute("SELECT * FROM doc_folders WHERE id = %s", (folder_id,)).fetchone()
    if not folder:
        abort(404)
    return folder

def _folder_breadcrumb(db, folder):
    """Remonte la chaîne des parents pour afficher le fil d'Ariane."""
    crumbs = []
    current = folder
    while current:
        crumbs.append(current)
        current = (
            db.execute("SELECT * FROM doc_folders WHERE id = %s", (current["parent_id"],)).fetchone()
            if current["parent_id"] else None
        )
    return list(reversed(crumbs))

def _human_size(size_bytes):
    if not size_bytes:
        return ""
    if size_bytes < 1024:
        return f"{size_bytes} o"
    if size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.0f} Ko"
    return f"{size_bytes / (1024 * 1024):.1f} Mo"

@app.route("/documents")
@app.route("/documents/dossier/<int:folder_id>")
def documents(folder_id=None):
    db = get_db()
    folder = _folder_or_404(db, folder_id) if folder_id else None
    breadcrumb = _folder_breadcrumb(db, folder) if folder else []
    if folder_id:
        subfolders = db.execute(
            "SELECT * FROM doc_folders WHERE parent_id = %s ORDER BY name", (folder_id,)
        ).fetchall()
        docs = db.execute(
            """SELECT id, folder_id, kind, filename, mimetype, size_bytes, url, title,
                      uploaded_by, uploaded_at, visibility, shared_group_id, shared_player_id
               FROM documents WHERE folder_id = %s
               ORDER BY uploaded_at DESC, id DESC""",
            (folder_id,),
        ).fetchall()
    else:
        subfolders = db.execute(
            "SELECT * FROM doc_folders WHERE parent_id IS NULL ORDER BY name"
        ).fetchall()
        docs = db.execute(
            """SELECT id, folder_id, kind, filename, mimetype, size_bytes, url, title,
                      uploaded_by, uploaded_at, visibility, shared_group_id, shared_player_id
               FROM documents WHERE folder_id IS NULL
               ORDER BY uploaded_at DESC, id DESC"""
        ).fetchall()
    # Nombre d'éléments par sous-dossier (dossiers enfants + documents) pour l'affichage.
    counts = {}
    for sub in subfolders:
        n_docs = db.execute(
            "SELECT COUNT(*) AS n FROM documents WHERE folder_id = %s", (sub["id"],)
        ).fetchone()["n"]
        n_dirs = db.execute(
            "SELECT COUNT(*) AS n FROM doc_folders WHERE parent_id = %s", (sub["id"],)
        ).fetchone()["n"]
        counts[sub["id"]] = n_docs + n_dirs
    groups = db.execute("SELECT id, name FROM player_groups ORDER BY name").fetchall()
    players = db.execute(
        "SELECT id, first_name, last_name FROM players ORDER BY last_name, first_name"
    ).fetchall()
    groups_by_id = {g["id"]: g["name"] for g in groups}
    players_by_id = {p["id"]: f"{p['first_name']} {p['last_name']}" for p in players}
    docs_view = []
    for d in docs:
        d = dict(d)
        d["icon"] = _doc_icon(d)
        d["ext"] = _doc_ext(d["filename"])
        d["inline"] = d["kind"] == "file" and d["ext"] in DOC_INLINE_EXTENSIONS
        d["size_human"] = _human_size(d["size_bytes"])
        d["can_delete"] = _doc_can_delete(d)
        d["date_human"] = (d["uploaded_at"] or "")[:10]
        if d["visibility"] == "players":
            d["sharing_label"] = "👥 Tous les joueurs"
        elif d["visibility"] == "group":
            d["sharing_label"] = f"👥 Groupe : {groups_by_id.get(d['shared_group_id'], '?')}"
        elif d["visibility"] == "player":
            d["sharing_label"] = f"👤 {players_by_id.get(d['shared_player_id'], '?')}"
        else:
            d["sharing_label"] = None
        docs_view.append(d)
    return render_template(
        "documents.html", folder=folder, breadcrumb=breadcrumb,
        subfolders=subfolders, folder_counts=counts, docs=docs_view,
        can_delete_folder={s["id"]: _doc_can_delete(dict(s)) for s in subfolders},
        player_groups=groups, players=players,
    )

@app.route("/documents/dossier", methods=["POST"])
def documents_create_folder():
    name = request.form.get("name", "").strip()
    parent_id = request.form.get("parent_id") or None
    if not name:
        flash("Merci d'indiquer un nom de dossier.", "error")
    else:
        db = get_db()
        db.execute(
            "INSERT INTO doc_folders (name, parent_id, created_by, created_at) VALUES (%s, %s, %s, %s)",
            (name, parent_id, session.get("user_email", ""), datetime.utcnow().isoformat()),
        )
        db.commit()
        flash(f"Dossier « {name} » créé.", "success")
    return redirect(url_for("documents", folder_id=parent_id) if parent_id else url_for("documents"))

def _read_sharing_fields(form):
    """Lit les 3 champs du formulaire de partage (documents.html) : 'visibility' vaut
    'staff' (par défaut, comportement historique — jamais visible aux joueurs), 'players'
    (tous les joueurs), 'group' (un groupe précis, avec shared_group_id) ou 'player' (un
    joueur précis, avec shared_player_id)."""
    visibility = form.get("visibility") or "staff"
    if visibility not in ("staff", "players", "group", "player"):
        visibility = "staff"
    shared_group_id = form.get("shared_group_id") or None
    shared_player_id = form.get("shared_player_id") or None
    if visibility != "group":
        shared_group_id = None
    if visibility != "player":
        shared_player_id = None
    return visibility, shared_group_id, shared_player_id

@app.route("/documents/upload", methods=["POST"])
def documents_upload():
    folder_id = request.form.get("folder_id") or None
    files = [f for f in request.files.getlist("files") if f and f.filename]
    if not files:
        flash("Merci de sélectionner au moins un fichier.", "error")
        return redirect(url_for("documents", folder_id=folder_id) if folder_id else url_for("documents"))
    visibility, shared_group_id, shared_player_id = _read_sharing_fields(request.form)
    db = get_db()
    saved = 0
    for file in files:
        payload = file.read()
        if not payload:
            continue
        db.execute(
            """INSERT INTO documents
               (folder_id, kind, filename, mimetype, size_bytes, data, uploaded_by, uploaded_at,
                visibility, shared_group_id, shared_player_id)
               VALUES (%s, 'file', %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (
                folder_id, file.filename,
                file.mimetype or "application/octet-stream",
                len(payload), psycopg2.Binary(payload),
                session.get("user_email", ""), datetime.utcnow().isoformat(),
                visibility, shared_group_id, shared_player_id,
            ),
        )
        saved += 1
    db.commit()
    flash(f"{saved} fichier{'s' if saved > 1 else ''} déposé{'s' if saved > 1 else ''}.", "success")
    return redirect(url_for("documents", folder_id=folder_id) if folder_id else url_for("documents"))

@app.route("/documents/lien", methods=["POST"])
def documents_add_link():
    folder_id = request.form.get("folder_id") or None
    title = request.form.get("title", "").strip()
    url = request.form.get("url", "").strip()
    if not url or not (url.startswith("http://") or url.startswith("https://")):
        flash("Merci de coller un lien complet (commençant par http:// ou https://).", "error")
    else:
        visibility, shared_group_id, shared_player_id = _read_sharing_fields(request.form)
        db = get_db()
        db.execute(
            """INSERT INTO documents (folder_id, kind, url, title, uploaded_by, uploaded_at,
                                       visibility, shared_group_id, shared_player_id)
               VALUES (%s, 'link', %s, %s, %s, %s, %s, %s, %s)""",
            (folder_id, url, title or url, session.get("user_email", ""), datetime.utcnow().isoformat(),
             visibility, shared_group_id, shared_player_id),
        )
        db.commit()
        flash("Lien ajouté.", "success")
    return redirect(url_for("documents", folder_id=folder_id) if folder_id else url_for("documents"))

def _serve_document(doc_id, inline):
    db = get_db()
    doc = db.execute("SELECT * FROM documents WHERE id = %s", (doc_id,)).fetchone()
    if not doc or doc["kind"] != "file":
        abort(404)
    data = bytes(doc["data"])
    disposition = "inline" if inline else "attachment"
    filename = (doc["filename"] or "document").replace('"', "")
    return Response(
        data,
        mimetype=doc["mimetype"] or "application/octet-stream",
        headers={
            "Content-Disposition": f'{disposition}; filename="{filename}"',
            "Content-Length": str(len(data)),
        },
    )

@app.route("/documents/<int:doc_id>/telecharger")
def documents_download(doc_id):
    return _serve_document(doc_id, inline=False)

@app.route("/documents/<int:doc_id>/apercu")
def documents_preview(doc_id):
    return _serve_document(doc_id, inline=True)

@app.route("/documents/<int:doc_id>/supprimer", methods=["POST"])
def documents_delete(doc_id):
    db = get_db()
    doc = db.execute(
        "SELECT id, folder_id, kind, filename, title, uploaded_by FROM documents WHERE id = %s",
        (doc_id,),
    ).fetchone()
    if not doc:
        abort(404)
    if not _doc_can_delete(dict(doc)):
        flash("Tu ne peux supprimer que tes propres dépôts (ou demande à l'admin).", "error")
    else:
        db.execute("DELETE FROM documents WHERE id = %s", (doc_id,))
        db.commit()
        flash("Document supprimé.", "success")
    folder_id = doc["folder_id"]
    return redirect(url_for("documents", folder_id=folder_id) if folder_id else url_for("documents"))

@app.route("/documents/dossier/<int:folder_id>/supprimer", methods=["POST"])
def documents_delete_folder(folder_id):
    db = get_db()
    folder = _folder_or_404(db, folder_id)
    parent_id = folder["parent_id"]
    if not _doc_can_delete(dict(folder)):
        flash("Tu ne peux supprimer que les dossiers que tu as créés (ou demande à l'admin).", "error")
        return redirect(url_for("documents", folder_id=folder_id))
    n_docs = db.execute(
        "SELECT COUNT(*) AS n FROM documents WHERE folder_id = %s", (folder_id,)
    ).fetchone()["n"]
    n_dirs = db.execute(
        "SELECT COUNT(*) AS n FROM doc_folders WHERE parent_id = %s", (folder_id,)
    ).fetchone()["n"]
    if (n_docs or n_dirs) and not session.get("is_admin"):
        flash("Ce dossier n'est pas vide : seul l'admin peut le supprimer avec son contenu.", "error")
        return redirect(url_for("documents", folder_id=folder_id))
    db.execute("DELETE FROM doc_folders WHERE id = %s", (folder_id,))
    db.commit()
    flash(f"Dossier « {folder['name']} » supprimé.", "success")
    return redirect(url_for("documents", folder_id=parent_id) if parent_id else url_for("documents"))

@app.route("/documents/dossier/<int:folder_id>/renommer", methods=["POST"])
def documents_rename_folder(folder_id):
    db = get_db()
    folder = _folder_or_404(db, folder_id)
    name = request.form.get("name", "").strip()
    if not name:
        flash("Merci d'indiquer un nom de dossier.", "error")
    elif not _doc_can_delete(dict(folder)):
        flash("Tu ne peux renommer que les dossiers que tu as créés (ou demande à l'admin).", "error")
    else:
        db.execute("UPDATE doc_folders SET name = %s WHERE id = %s", (name, folder_id))
        db.commit()
        flash("Dossier renommé.", "success")
    return redirect(url_for("documents", folder_id=folder_id))

# ---------------------------------------------------------------------------
# ESPACE JOUEUR — un joueur connecté n'a accès qu'à 3 pages : son planning (le
# calendrier du staff, en lecture seule), les documents qui lui sont partagés
# (par lui-même, par son groupe, ou par le staff pour tous les joueurs), et ses
# propres statistiques individuelles cumulées sur la saison. Voir
# PLAYER_ALLOWED_ENDPOINTS / gate_player_access() en haut du fichier pour le
# garde-fou qui empêche l'accès à tout le reste du site.
# ---------------------------------------------------------------------------
def _normalize_name(s):
    """Nettoie un nom pour une comparaison automatique fiable entre l'orthographe du
    fichier Excel du club et celle tapée dans Sportscode (accents, casse, tirets,
    espaces peuvent différer légèrement) : enlève les accents, passe en minuscules, ne
    garde que lettres et chiffres."""
    normalized = unicodedata.normalize("NFKD", s or "")
    normalized = "".join(c for c in normalized if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]", "", normalized.lower())

def _current_player(db):
    """Renvoie la ligne 'players' du joueur actuellement connecté, ou None (compte
    staff/admin, ou session invalide)."""
    if not session.get("is_player"):
        return None
    return db.execute("SELECT * FROM players WHERE id = %s", (session.get("player_id"),)).fetchone()

def _doc_visible_to_player(doc, player):
    """Un document n'est visible à un joueur que si le staff l'a explicitement partagé
    avec lui : avec tous les joueurs, avec son groupe, ou avec lui précisément. Jamais
    les documents restés en visibilité 'staff' (comportement par défaut, historique)."""
    vis = doc.get("visibility") or "staff"
    if vis == "players":
        return True
    if vis == "group":
        return player.get("group_id") is not None and doc.get("shared_group_id") == player.get("group_id")
    if vis == "player":
        return doc.get("shared_player_id") == player.get("id")
    return False

def _match_player_stats_name(player, instances):
    """Correspondance automatique stricte entre le nom du joueur (fichier Excel du club)
    et le nom codé dans Sportscode (qui ne contient que le NOM DE FAMILLE, voir
    SQUAD_ROSTER dans parser.py). On compare les noms normalisés (sans accents/casse/
    ponctuation) : si le nom de famille du joueur correspond à un nom codé dans les
    matchs de la saison, on l'utilise. Sinon, la page affiche un message plutôt que des
    statistiques erronées. NB : 2 joueurs qui partagent exactement le même nom de famille
    ne peuvent pas être distingués automatiquement par ce système (rare, mais possible)."""
    attack = compute_player_attack_table(instances)
    defense = compute_player_defense_table(instances)
    ruck = compute_player_ruck_table(instances)
    coded_names = {r["name"] for r in attack["rows"]} | {r["name"] for r in defense["rows"]} | {r["name"] for r in ruck["rows"]}
    target = _normalize_name(player["last_name"])
    for coded in coded_names:
        if _normalize_name(coded) == target:
            return coded
    return None

@app.route("/mon-espace")
def player_home():
    db = get_db()
    player = _current_player(db)
    if not player:
        abort(404)
    return render_template("player_home.html", player=player)

@app.route("/mes-documents")
def player_documents():
    db = get_db()
    player = _current_player(db)
    if not player:
        abort(404)
    docs = db.execute(
        """SELECT id, kind, filename, mimetype, size_bytes, url, title, uploaded_by, uploaded_at
           FROM documents
           WHERE visibility = 'players'
              OR (visibility = 'group' AND shared_group_id = %s)
              OR (visibility = 'player' AND shared_player_id = %s)
           ORDER BY uploaded_at DESC, id DESC""",
        (player["group_id"], player["id"]),
    ).fetchall()
    docs_view = []
    for d in docs:
        d = dict(d)
        d["icon"] = _doc_icon(d)
        d["ext"] = _doc_ext(d["filename"])
        d["inline"] = d["kind"] == "file" and d["ext"] in DOC_INLINE_EXTENSIONS
        d["size_human"] = _human_size(d["size_bytes"])
        d["date_human"] = (d["uploaded_at"] or "")[:10]
        docs_view.append(d)
    return render_template("player_documents.html", docs=docs_view, player=player)

def _serve_player_document(doc_id, inline):
    db = get_db()
    player = _current_player(db)
    if not player:
        abort(404)
    doc = db.execute("SELECT * FROM documents WHERE id = %s", (doc_id,)).fetchone()
    if not doc or doc["kind"] != "file" or not _doc_visible_to_player(dict(doc), dict(player)):
        abort(404)
    data = bytes(doc["data"])
    disposition = "inline" if inline else "attachment"
    filename = (doc["filename"] or "document").replace('"', "")
    return Response(
        data,
        mimetype=doc["mimetype"] or "application/octet-stream",
        headers={
            "Content-Disposition": f'{disposition}; filename="{filename}"',
            "Content-Length": str(len(data)),
        },
    )

@app.route("/mes-documents/<int:doc_id>/telecharger")
def player_document_download(doc_id):
    return _serve_player_document(doc_id, inline=False)

@app.route("/mes-documents/<int:doc_id>/apercu")
def player_document_preview(doc_id):
    return _serve_player_document(doc_id, inline=True)

@app.route("/mes-stats")
def player_stats():
    db = get_db()
    player = _current_player(db)
    if not player:
        abort(404)
    matches_with_instances, selected, selected_ids, qs = _season_context()
    instances = _season_instances(selected)
    matched_name = _match_player_stats_name(player, instances) if instances else None
    attack_row = defense_row = ruck_row = None
    matches_played = 0
    if matched_name:
        comparison = compute_player_comparison(instances, matched_name, matched_name)
        attack_row = comparison["attack_rows"][0]
        defense_row = comparison["defense_rows"][0]
        ruck_row = comparison["ruck_rows"][0]
        matches_played = sum(
            1 for m in selected
            if any(i["kind"] == "player" and i["code_raw"] == matched_name for i in m["instances"])
        )
    return render_template(
        "player_stats.html", player=player, matched_name=matched_name,
        attack=attack_row, defense=defense_row, ruck=ruck_row,
        matches_played=matches_played, total_matches=len(matches_with_instances),
    )

@app.route("/mes-evaluations")
def player_evaluations():
    """P.P.I.D du joueur connecté : historique rugby + physique renseigné par le staff, et
    formulaire pour sa propre auto-évaluation physique (seule partie modifiable côté
    joueur — voir player_ppid_auto_update). Reprend numériquement ce que le club envoyait
    jusque-là par WhatsApp/Hudl."""
    db = get_db()
    player = _current_player(db)
    if not player:
        abort(404)
    rugby_evals = [
        _ppid_rugby_row_view(r) for r in db.execute(
            "SELECT * FROM ppid_rugby_evals WHERE player_id = %s ORDER BY eval_date DESC NULLS LAST, id DESC",
            (player["id"],),
        ).fetchall()
    ]
    physical_evals = [
        _ppid_physical_row_view(r) for r in db.execute(
            "SELECT * FROM ppid_physical_evals WHERE player_id = %s ORDER BY eval_date DESC NULLS LAST, id DESC",
            (player["id"],),
        ).fetchall()
    ]
    entretiens = [
        _ppid_entretien_row_view(r) for r in db.execute(
            "SELECT * FROM ppid_entretiens WHERE player_id = %s ORDER BY entretien_date DESC, id DESC",
            (player["id"],),
        ).fetchall()
    ]
    _ppid_compute_trends(rugby_evals, PPID_RUGBY_CATEGORY_KEYS, lambda r: r.get("note"))
    _ppid_compute_trends(physical_evals, [key for key, _label in PPID_PHYSICAL_CATEGORIES], lambda r: r.get("coach"))
    return render_template(
        "player_evaluations.html", player=player, rugby_evals=rugby_evals, physical_evals=physical_evals,
        entretiens=entretiens, ppid_timeline=_ppid_timeline(rugby_evals, physical_evals, entretiens),
        ppid_rugby_categories=_ppid_rugby_categories_for_position(player.get("ppid_position")),
        ppid_physical_categories=PPID_PHYSICAL_CATEGORIES,
        ppid_physical_notes=PPID_PHYSICAL_NOTES,
        ppid_profil_rows=PPID_PROFIL_ROWS,
        ppid_profil_par_poste=PPID_PROFIL_PAR_POSTE,
    )

@app.route("/mes-evaluations/physique/<int:eval_id>/auto", methods=["POST"])
def player_ppid_auto_update(eval_id):
    db = get_db()
    player = _current_player(db)
    if not player:
        abort(404)
    row = db.execute(
        "SELECT * FROM ppid_physical_evals WHERE id = %s AND player_id = %s", (eval_id, player["id"]),
    ).fetchone()
    if not row:
        abort(404)
    ratings = _ppid_physical_auto_ratings_from_form(request.form, row["ratings"])
    db.execute(
        "UPDATE ppid_physical_evals SET ratings = %s, updated_at = %s WHERE id = %s",
        (ratings, datetime.utcnow().isoformat(), eval_id),
    )
    db.commit()
    flash("Ton auto-évaluation a été enregistrée.", "success")
    return redirect(url_for("player_evaluations"))

# ---------------------------------------------------------------------------
# ANALYSE VIDÉO — page d'entrée du pôle qui regroupe tout ce qui existait sur
# le site avant l'ajout des espaces Documents / Calendrier / Cahier des charges
# (Matchs, Bilan de saison, Adversaires, Tendances, Prochain match, Effectif).
# ---------------------------------------------------------------------------
@app.route("/analyse-video")
def analyse_video():
    return render_template("analyse_video.html")

# ---------------------------------------------------------------------------
# CALENDRIER — agenda d'événements du staff (réunions, déplacements, rendez-vous
# médicaux...), distinct des matchs/entraînements gérés ailleurs sur le site.
# Règles identiques à l'espace Documents : tout le staff ajoute, chacun gère ses
# propres événements, l'admin gère tout ; invisible en mode démo.
# ---------------------------------------------------------------------------
def _cal_can_edit(row):
    """Réutilise la même règle que les documents : admin = tout, sinon = ses propres
    événements uniquement (comparaison sur created_by)."""
    return _doc_can_delete(dict(row))

def _cal_event_json(row):
    r = dict(row)
    return {
        "id": r["id"],
        "title": r["title"],
        "description": r.get("description") or "",
        "event_date": r["event_date"],
        "event_time": r.get("event_time") or "",
        "created_by": r.get("created_by") or "",
        "can_edit": _cal_can_edit(r),
    }

@app.route("/calendrier")
def calendrier():
    # La page est rendue côté client (vue Jour / Semaine / Mois interactive, façon iPhone) :
    # ce endpoint ne fait que servir le squelette HTML + la date du jour. Les événements
    # sont chargés en JSON via /calendrier/api/events selon la période affichée.
    today = datetime.utcnow().strftime("%Y-%m-%d")
    return render_template("calendrier.html", today=today)

@app.route("/calendrier/api/events")
def calendrier_api_events():
    """Renvoie les événements dans une période [start, end] (AAAA-MM-JJ, bornes incluses).
    Sans paramètres, renvoie tout (utilisé en secours)."""
    start = request.args.get("start", "").strip()
    end = request.args.get("end", "").strip()
    db = get_db()
    if start and end:
        rows = db.execute(
            "SELECT * FROM calendar_events WHERE event_date >= %s AND event_date <= %s "
            "ORDER BY event_date ASC, COALESCE(event_time, '99:99') ASC, id ASC",
            (start, end),
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM calendar_events ORDER BY event_date ASC, COALESCE(event_time, '99:99') ASC, id ASC"
        ).fetchall()
    return jsonify({"events": [_cal_event_json(r) for r in rows]})

@app.route("/calendrier/api/ajouter", methods=["POST"])
def calendrier_api_add():
    data = request.get_json(silent=True) or request.form
    title = (data.get("title") or "").strip()
    event_date = (data.get("event_date") or "").strip()
    event_time = (data.get("event_time") or "").strip()
    description = (data.get("description") or "").strip()
    if not title or not event_date:
        return jsonify({"error": "Merci d'indiquer au moins un titre et une date."}), 400
    db = get_db()
    row = db.execute(
        """INSERT INTO calendar_events (title, description, event_date, event_time, created_by, created_at)
           VALUES (%s, %s, %s, %s, %s, %s) RETURNING *""",
        (title, description or None, event_date, event_time or None,
         session.get("user_email", ""), datetime.utcnow().isoformat()),
    ).fetchone()
    db.commit()
    return jsonify({"event": _cal_event_json(row)})

@app.route("/calendrier/api/<int:event_id>/modifier", methods=["POST"])
def calendrier_api_edit(event_id):
    db = get_db()
    row = db.execute("SELECT * FROM calendar_events WHERE id = %s", (event_id,)).fetchone()
    if not row:
        return jsonify({"error": "Introuvable."}), 404
    if not _cal_can_edit(row):
        return jsonify({"error": "Tu ne peux modifier que tes propres événements (ou demande à l'admin)."}), 403
    data = request.get_json(silent=True) or request.form
    title = (data.get("title") or "").strip()
    event_date = (data.get("event_date") or "").strip()
    event_time = (data.get("event_time") or "").strip()
    description = (data.get("description") or "").strip()
    if not title or not event_date:
        return jsonify({"error": "Merci d'indiquer au moins un titre et une date."}), 400
    updated = db.execute(
        """UPDATE calendar_events SET title = %s, description = %s, event_date = %s, event_time = %s
           WHERE id = %s RETURNING *""",
        (title, description or None, event_date, event_time or None, event_id),
    ).fetchone()
    db.commit()
    return jsonify({"event": _cal_event_json(updated)})

@app.route("/calendrier/api/<int:event_id>/supprimer", methods=["POST"])
def calendrier_api_delete(event_id):
    db = get_db()
    row = db.execute("SELECT * FROM calendar_events WHERE id = %s", (event_id,)).fetchone()
    if not row:
        return jsonify({"error": "Introuvable."}), 404
    if not _cal_can_edit(row):
        return jsonify({"error": "Tu ne peux supprimer que tes propres événements (ou demande à l'admin)."}), 403
    db.execute("DELETE FROM calendar_events WHERE id = %s", (event_id,))
    db.commit()
    return jsonify({"ok": True})

# ---------------------------------------------------------------------------
# CAHIER DES CHARGES — suivi de tâches/exigences du staff, en 3 colonnes
# (à faire / en cours / fait). Tout le staff peut créer une tâche et déplacer
# n'importe quelle tâche d'une colonne à l'autre (travail collaboratif) ; seule
# la personne qui l'a créée (ou l'admin) peut la supprimer. Invisible en démo.
# ---------------------------------------------------------------------------
CHARGES_STATUSES = ["a_faire", "en_cours", "fait"]
CHARGES_STATUS_LABELS = {"a_faire": "À faire", "en_cours": "En cours", "fait": "Fait"}

# ---------------------------------------------------------------------------
# P.P.I.D — reprise numérique, plus lisible et visuelle, du cahier d'entraînement
# individuel papier du club (profil par poste / évaluation rugby / évaluation
# physique), intégrée dans l'onglet de chaque joueur au Cahier des charges.
# ---------------------------------------------------------------------------
PPID_POSITIONS = ["Pilier", "Talon", "2L", "3LC", "3LA", "9", "10", "12-13", "11-14", "15"]

# Clés techniques des 9 critères d'évaluation rugby (stockage en base, stables dans le
# temps) : les 8 premières correspondent 1-pour-1 aux 8 lignes de PPID_PROFIL_PAR_POSTE
# (même ordre), la 9ème ("dureté / état d'esprit") est générique et hors grille de poste.
# Les LIBELLÉS affichés, eux, ne sont plus fixes : voir _ppid_rugby_categories_for_position().
PPID_RUGBY_CATEGORY_KEYS = [
    "melee_fermee", "technique_touche", "plaquer_contest", "soutenir_rucker", "duel_off",
    "habilite_technique", "se_deplacer", "comprehension_systeme", "durete_etat_esprit",
]
PPID_RUGBY_NOTES = ["MOY", "BIEN", "EXL"]

PPID_PHYSICAL_CATEGORIES = [
    ("capacite_entrainer", "Capacité à s'entraîner / Rustisité"),
    ("poids_composition", "Poids / Composition corporelle"),
    ("force", "Force"),
    ("puissance", "Puissance"),
    ("vitesse", "Vitesse"),
    ("conditionning", "Conditionning / Énergétique"),
]
PPID_PHYSICAL_NOTES = ["Moyen", "Bien", "Excellent"]
# Rang numérique des notes des 2 échelles (rugby ET physique), pour calculer les flèches de
# progression d'un critère entre 2 évaluations successives — voir _ppid_compute_trends().
PPID_NOTE_RANK = {"MOY": 1, "BIEN": 2, "EXL": 3, "Moyen": 1, "Bien": 2, "Excellent": 3}

PPID_ENTRETIEN_TYPES = ["Briefing", "Retour de match", "Préparation de match", "Entretien individuel", "Autre"]

# Tableau de référence "Profil par poste" du cahier papier : les critères
# d'auto-évaluation attendus, déclinés par poste. Contenu générique du club, identique
# pour tout le monde — seule la colonne du poste PPID du joueur sélectionné est mise en
# évidence (voir cahier_charges.html). Transcrit depuis le cahier papier du club ; si une
# case ne correspond pas exactement à l'original, elle se corrige en un message.
PPID_PROFIL_PAR_POSTE = {
    "Mêlée fermée": {
        "Pilier": "Mêlée fermée", "Talon": "Lancer", "2L": "Sauter / lifter / touche / CE et CR",
        "3LC": "Mêlée : gestion du ballon", "3LA": "Soutenir / rucker", "9": "Transmission",
        "10": "Lecture", "12-13": "Lecture off", "11-14": "Duel / franchissement",
        "15": "Gestion du 3ème rideau / CA",
    },
    "Technique de lift sur touche, coup d'envoi et coup de renvoi": {
        "Pilier": "Technique de lift sur touche, coup d'envoi et coup de renvoi", "Talon": "Mêlée fermée",
        "2L": "Mêlée fermée", "3LC": "Assure la continuité", "3LA": "Plaquer / contest / défendre",
        "9": "Colle au ballon", "10": "Stratégie", "12-13": "Duel off / franchissement",
        "11-14": "Gestion des contre-attaques", "15": "Compréhension système",
    },
    "Plaquer / Contest": {
        "Pilier": "Plaquer / contest", "Talon": "Plaquer / contest / défendre", "2L": "Contest / plaquer",
        "3LC": "Gagner les contacts / jouer les duels / franchir", "3LA": "Plaquage / circulation défensive",
        "9": "Circulation déf. / plaquer / contest / défendre", "10": "Lecture défense",
        "12-13": "1 contre 1 déf.", "11-14": "Plaquer / contest", "15": "Duel / franchissement",
    },
    "Soutenir / Rucker": {
        "Pilier": "Soutenir / rucker", "Talon": "Soutenir / rucker", "2L": "Défense de maul",
        "3LC": "Sauter / lifter / lecture touche", "3LA": "Assure la continuité", "9": "Stratégie",
        "10": "Transmission", "12-13": "Habileté technique (main / pied)", "11-14": "Soutenir / ruck aérien",
        "15": "Duel / franchissement",
    },
    "Duel off": {
        "Pilier": "Duel off", "Talon": "Défense de maul", "2L": "Sauter / lifter / lecture touche",
        "3LC": "Assure la continuité", "3LA": "Duel off",
        "9": "Jeu au pied / sortie de camp et pression", "10": "Jeu au pied (pression / occupation / CE / CR / drop)",
        "12-13": "Communication", "11-14": "Soutenir / ruck offensif", "15": "Habileté technique (mains / pied)",
    },
    "Habileté technique": {
        "Pilier": "Habileté technique", "Talon": "Duel off / porteur de balle", "2L": "Duel off / franchissement",
        "3LC": "Habileté technique", "3LA": "Se déplacer / enchaîner les tâches", "9": "Habileté technique",
        "10": "Gestion contre-attaque", "12-13": "Se déplacer / enchaîner les tâches",
        "11-14": "Habileté technique (mains / pied)", "15": "Réception / duel aérien",
    },
    "Se déplacer / Enchaîner les actions": {
        "Pilier": "Se déplacer / enchaîner les tâches", "Talon": "Se déplacer / enchaîner les tâches",
        "2L": "Habileté technique", "3LC": "Se déplacer / enchaîner les tâches", "3LA": "Leadership",
        "9": "Gestion contre-attaque", "10": "Se déplacer / enchaîner les tâches",
        "12-13": "Compréhension système", "11-14": "Compréhension système", "15": "Se déplacer / enchaîner les actions",
    },
    "Compréhension système": {
        "Pilier": "Compréhension système", "Talon": "Compréhension système", "2L": "Compréhension système",
        "3LC": "Se déplacer / enchaîner les tâches", "3LA": "Compréhension système", "9": "Soutenir / rucker",
        "10": "Se déplacer / enchaîner les tâches", "12-13": "Compréhension système",
        "11-14": "Compréhension système", "15": "Se déplacer / enchaîner les actions",
    },
}
PPID_PROFIL_ROWS = list(PPID_PROFIL_PAR_POSTE.keys())

def _ppid_rugby_categories_for_position(position):
    """Personnalise les libellés des 9 critères d'évaluation rugby selon le poste PPID du
    joueur : au lieu d'imposer les mêmes 9 intitulés génériques à tout le monde (illisible
    quand on doit remplir 59 fiches), chaque poste voit directement SES critères, repris de
    sa colonne dans le tableau de référence « Profil par poste » (ex. un 9 voit « Transmission »
    là où un Pilier voit « Mêlée fermée » — même case techniquement, sens différent pour son
    poste). Sans poste PPID réglé pour ce joueur, on retombe sur les intitulés génériques
    (les lignes du tableau de référence) plutôt que de deviner."""
    categories = []
    for key, row_label in zip(PPID_RUGBY_CATEGORY_KEYS, PPID_PROFIL_ROWS):
        if position and position in PPID_POSITIONS:
            label = PPID_PROFIL_PAR_POSTE[row_label].get(position) or row_label
        else:
            label = row_label
        categories.append((key, label))
    categories.append(("durete_etat_esprit", "Dureté / État d'esprit"))
    return categories

def _ppid_ratings_view(raw, category_keys):
    """Décode le JSON de notes stocké en texte, en garantissant une entrée pour chaque
    catégorie connue (au cas où de nouvelles catégories seraient ajoutées après coup)."""
    try:
        data = json.loads(raw) if raw else {}
    except (TypeError, ValueError):
        data = {}
    return {key: data.get(key) or {} for key in category_keys}

PPID_MONTHS_FR_FULL = [
    "janvier", "février", "mars", "avril", "mai", "juin",
    "juillet", "août", "septembre", "octobre", "novembre", "décembre",
]

def _ppid_date_human(date_str):
    """Formate une date ISO ('2026-08-15') en date lisible en français ('15 août 2026') —
    demande du manager pour que les dates des suivis PPID soient claires d'un coup d'œil sur
    les cartes, plutôt que le format ISO brut (peu lisible, ambigu jour/mois pour un lecteur
    non technique). Retourne la valeur d'origine telle quelle si elle ne ressemble pas à une
    date ISO complète, pour ne jamais planter ni afficher n'importe quoi sur une donnée
    inattendue."""
    if not date_str or len(date_str) < 10:
        return date_str
    try:
        year, month, day = int(date_str[0:4]), int(date_str[5:7]), int(date_str[8:10])
    except ValueError:
        return date_str
    if not (1 <= month <= 12):
        return date_str
    return f"{day} {PPID_MONTHS_FR_FULL[month - 1]} {year}"

def _ppid_rugby_row_view(row):
    r = dict(row)
    r["ratings"] = _ppid_ratings_view(r.get("ratings"), PPID_RUGBY_CATEGORY_KEYS)
    r["date_human"] = _ppid_date_human(r.get("eval_date") or (r.get("created_at") or "")[:10])
    return r

def _ppid_physical_row_view(row):
    r = dict(row)
    r["ratings"] = _ppid_ratings_view(r.get("ratings"), [key for key, _label in PPID_PHYSICAL_CATEGORIES])
    r["date_human"] = _ppid_date_human(r.get("eval_date") or (r.get("created_at") or "")[:10])
    return r

def _ppid_entretien_row_view(row):
    r = dict(row)
    r["date_human"] = _ppid_date_human(r.get("entretien_date") or (r.get("created_at") or "")[:10])
    return r

def _ppid_compute_trends(evals, category_keys, value_getter):
    """La « flèche du temps » demandée par le manager : pour chaque évaluation d'une série
    (rugby OU physique, pour un même joueur), calcule la tendance de chaque critère par
    rapport à l'évaluation précédente de LA MÊME série — on ne lit plus chaque évaluation
    isolément, la progression du joueur dans la durée saute aux yeux. Écrit le résultat dans
    eval['trends'][key] = 'up' | 'down' | 'flat' | None (None = pas de valeur précédente
    comparable, typiquement la toute première évaluation du joueur sur ce critère). Modifie
    les dicts en place : rugby_evals/physical_evals sont ensuite réutilisés tels quels par
    _ppid_timeline, donc les tendances y sont automatiquement disponibles aussi."""
    chronological = sorted(evals, key=lambda e: e.get("eval_date") or (e.get("created_at") or ""))
    last_rank = {}
    for ev in chronological:
        trends = {}
        for key in category_keys:
            rank = PPID_NOTE_RANK.get(value_getter(ev["ratings"].get(key) or {}))
            prev = last_rank.get(key)
            if rank is not None and prev is not None:
                trends[key] = "up" if rank > prev else ("down" if rank < prev else "flat")
            else:
                trends[key] = None
            if rank is not None:
                last_rank[key] = rank
        ev["trends"] = trends

def _ppid_bilan_stats(evals, category_keys, value_getter):
    """Petit résumé d'une série d'évaluations (rugby OU physique) pour le bilan global du
    joueur : nombre total de bilans, le plus récent, la répartition des notes sur CE dernier
    bilan (photo du niveau actuel du joueur, sans avoir à relire chaque critère un par un), et
    — s'il existe un bilan précédent pour comparer — le nombre de critères en progression /
    stable / en recul (réutilise les tendances déjà calculées par _ppid_compute_trends, donc
    aucun recalcul). evals doit déjà être trié du plus récent au plus ancien."""
    if not evals:
        return {"count": 0, "latest": None, "note_counts": None, "trend_counts": None}
    latest = evals[0]
    note_counts = {"exl": 0, "bien": 0, "moy": 0}
    for key in category_keys:
        rank = PPID_NOTE_RANK.get(value_getter(latest["ratings"].get(key) or {}))
        if rank == 3:
            note_counts["exl"] += 1
        elif rank == 2:
            note_counts["bien"] += 1
        elif rank == 1:
            note_counts["moy"] += 1
    trends = latest.get("trends") or {}
    trend_counts = None
    if any(v is not None for v in trends.values()):
        trend_counts = {"up": 0, "down": 0, "flat": 0}
        for v in trends.values():
            if v in trend_counts:
                trend_counts[v] += 1
    return {"count": len(evals), "latest": latest, "note_counts": note_counts, "trend_counts": trend_counts}

def _ppid_bilan_global(rugby_evals, physical_evals, entretiens):
    """Bilan global affiché en haut du Cahier des charges (sous les critères clés) : un
    résumé compact du joueur sur ses 3 suivis PPID (rugby / physique / entretiens) en un coup
    d'œil, plutôt que d'avoir à parcourir toute la timeline pour se faire une idée. rugby_evals
    et physical_evals doivent déjà porter leurs 'trends' (voir _ppid_compute_trends), et les 3
    listes doivent déjà être triées du plus récent au plus ancien (ordre des requêtes SQL)."""
    return {
        "rugby": _ppid_bilan_stats(rugby_evals, PPID_RUGBY_CATEGORY_KEYS, lambda r: r.get("note")),
        "physical": _ppid_bilan_stats(
            physical_evals, [key for key, _label in PPID_PHYSICAL_CATEGORIES], lambda r: r.get("coach")
        ),
        "entretien": {"count": len(entretiens), "latest": entretiens[0] if entretiens else None},
    }

PPID_MONTHS_FR_ABBR = [
    "Janv.", "Févr.", "Mars", "Avr.", "Mai", "Juin",
    "Juil.", "Août", "Sept.", "Oct.", "Nov.", "Déc.",
]

def _ppid_month_label(date_str):
    """Convertit une date ISO ('2026-08-15' ou '2026-08-15T10:00:00') en repère de mois
    court pour l'axe temporel ('Août 26') — retourne None si la date est absente/invalide
    (l'axe n'affiche alors pas de repère pour cette entrée plutôt que de deviner)."""
    if not date_str or len(date_str) < 7:
        return None
    try:
        year, month = int(date_str[0:4]), int(date_str[5:7])
    except ValueError:
        return None
    if not (1 <= month <= 12):
        return None
    return f"{PPID_MONTHS_FR_ABBR[month - 1]} {year % 100:02d}"

def _ppid_timeline(rugby_evals, physical_evals, entretiens):
    """Fusionne les 3 flux du P.P.I.D (évaluation rugby, évaluation physique, cahier
    d'entretien) en une seule chronologie triée par date décroissante — pour voir d'un
    coup tout ce qui concerne un joueur plutôt que de parcourir 3 sections séparées.
    Chaque entrée porte aussi son repère de mois ('month_label', ex. 'Août 26') pour l'axe
    temporel gradué par mois demandé par le manager. On calcule le repère sur CHAQUE entrée
    (pas seulement la première du mois) : c'est le JS côté template qui décide, au moment de
    l'affichage, sur quelle carte le faire apparaître — nécessaire pour que le regroupement
    par mois reste correct même quand les filtres par type masquent certaines entrées."""
    items = []
    for ev in rugby_evals:
        items.append({"type": "rugby", "sort_key": ev.get("eval_date") or (ev.get("created_at") or ""), "data": ev})
    for ev in physical_evals:
        items.append({"type": "physical", "sort_key": ev.get("eval_date") or (ev.get("created_at") or ""), "data": ev})
    for e in entretiens:
        items.append({"type": "entretien", "sort_key": e.get("entretien_date") or (e.get("created_at") or ""), "data": e})
    items.sort(key=lambda it: it["sort_key"] or "", reverse=True)
    for it in items:
        it["month_label"] = _ppid_month_label(it["sort_key"])
    return items

def _ppid_rugby_ratings_from_form(form):
    ratings = {}
    for key in PPID_RUGBY_CATEGORY_KEYS:
        note = (form.get(f"note__{key}") or "").strip()
        commentaire = (form.get(f"commentaire__{key}") or "").strip()
        if note or commentaire:
            ratings[key] = {"note": note or None, "commentaire": commentaire or None}
    return json.dumps(ratings, ensure_ascii=False)

def _ppid_physical_coach_ratings_from_form(form, existing_raw):
    """Fusionne les notes 'coach' saisies par le staff avec les 'auto' déjà présentes
    (remplies par le joueur) : le formulaire staff ne doit jamais écraser l'auto-évaluation
    du joueur, et inversement (voir ppid_physical_auto_update)."""
    existing = _ppid_ratings_view(existing_raw, PPID_PHYSICAL_CATEGORIES)
    ratings = {}
    for key, _label in PPID_PHYSICAL_CATEGORIES:
        coach = (form.get(f"coach__{key}") or "").strip()
        entry = dict(existing.get(key) or {})
        if coach:
            entry["coach"] = coach
        elif "coach" in entry:
            del entry["coach"]
        if entry:
            ratings[key] = entry
    return json.dumps(ratings, ensure_ascii=False)

def _ppid_physical_auto_ratings_from_form(form, existing_raw):
    existing = _ppid_ratings_view(existing_raw, PPID_PHYSICAL_CATEGORIES)
    ratings = {}
    for key, _label in PPID_PHYSICAL_CATEGORIES:
        auto = (form.get(f"auto__{key}") or "").strip()
        entry = dict(existing.get(key) or {})
        if auto:
            entry["auto"] = auto
        elif "auto" in entry:
            del entry["auto"]
        if entry:
            ratings[key] = entry
    return json.dumps(ratings, ensure_ascii=False)

@app.route("/cahier-des-charges")
def cahier_charges():
    db = get_db()
    joueur_id = request.args.get("joueur", type=int)
    players = db.execute(
        """SELECT p.*, g.name AS group_name FROM players p
           LEFT JOIN player_groups g ON g.id = p.group_id
           ORDER BY g.name NULLS LAST, p.last_name, p.first_name"""
    ).fetchall()
    # Regroupés par groupe (Avants/Trois-quarts...) pour une navigation plus lisible que
    # 59 joueurs à plat — voir cahier_charges.html (barre de recherche + sections repliables).
    grouped_players = {}
    for p in players:
        grouped_players.setdefault(p["group_name"] or "Sans groupe", []).append(p)
    selected_player = None
    if joueur_id:
        selected_player = db.execute(
            """SELECT p.*, g.name AS group_name FROM players p
               LEFT JOIN player_groups g ON g.id = p.group_id
               WHERE p.id = %s""",
            (joueur_id,),
        ).fetchone()
        if not selected_player:
            joueur_id = None
    # IS NOT DISTINCT FROM plutôt que '=' : gère nativement le cas joueur_id=None (tâches
    # générales du staff, sans joueur) sans avoir à écrire 2 requêtes différentes.
    rows = db.execute(
        "SELECT * FROM charges_items WHERE player_id IS NOT DISTINCT FROM %s ORDER BY created_at DESC, id DESC",
        (joueur_id,),
    ).fetchall()
    columns = {status: [] for status in CHARGES_STATUSES}
    for r in rows:
        r = dict(r)
        r["can_delete"] = _doc_can_delete(r)
        r["date_human"] = (r["created_at"] or "")[:10]
        columns.setdefault(r["status"], []).append(r)
    docs_view = []
    rugby_evals, physical_evals, entretiens = [], [], []
    if selected_player:
        docs = db.execute(
            """SELECT * FROM documents WHERE shared_player_id = %s AND visibility = 'player'
               ORDER BY uploaded_at DESC, id DESC""",
            (joueur_id,),
        ).fetchall()
        for d in docs:
            d = dict(d)
            d["icon"] = _doc_icon(d)
            d["ext"] = _doc_ext(d["filename"])
            d["inline"] = d["kind"] == "file" and d["ext"] in DOC_INLINE_EXTENSIONS
            d["size_human"] = _human_size(d["size_bytes"])
            d["can_delete"] = _doc_can_delete(d)
            d["date_human"] = (d["uploaded_at"] or "")[:10]
            docs_view.append(d)
        rugby_evals = [
            _ppid_rugby_row_view(r) for r in db.execute(
                "SELECT * FROM ppid_rugby_evals WHERE player_id = %s ORDER BY eval_date DESC NULLS LAST, id DESC",
                (joueur_id,),
            ).fetchall()
        ]
        physical_evals = [
            _ppid_physical_row_view(r) for r in db.execute(
                "SELECT * FROM ppid_physical_evals WHERE player_id = %s ORDER BY eval_date DESC NULLS LAST, id DESC",
                (joueur_id,),
            ).fetchall()
        ]
        entretiens = [
            _ppid_entretien_row_view(r) for r in db.execute(
                "SELECT * FROM ppid_entretiens WHERE player_id = %s ORDER BY entretien_date DESC, id DESC",
                (joueur_id,),
            ).fetchall()
        ]
        _ppid_compute_trends(rugby_evals, PPID_RUGBY_CATEGORY_KEYS, lambda r: r.get("note"))
        _ppid_compute_trends(physical_evals, [key for key, _label in PPID_PHYSICAL_CATEGORIES], lambda r: r.get("coach"))
    return render_template(
        "cahier_charges.html", columns=columns, statuses=CHARGES_STATUSES,
        status_labels=CHARGES_STATUS_LABELS, total_count=len(rows),
        grouped_players=grouped_players, selected_player=selected_player, joueur_id=joueur_id, docs=docs_view,
        rugby_evals=rugby_evals, physical_evals=physical_evals, entretiens=entretiens,
        ppid_timeline=_ppid_timeline(rugby_evals, physical_evals, entretiens),
        ppid_bilan=_ppid_bilan_global(rugby_evals, physical_evals, entretiens) if selected_player else None,
        ppid_positions=PPID_POSITIONS,
        ppid_rugby_categories=_ppid_rugby_categories_for_position(selected_player.get("ppid_position") if selected_player else None),
        ppid_rugby_notes=PPID_RUGBY_NOTES, ppid_physical_categories=PPID_PHYSICAL_CATEGORIES,
        ppid_physical_notes=PPID_PHYSICAL_NOTES, ppid_profil_rows=PPID_PROFIL_ROWS,
        ppid_profil_par_poste=PPID_PROFIL_PAR_POSTE, ppid_entretien_types=PPID_ENTRETIEN_TYPES,
    )

@app.route("/cahier-des-charges/ajouter", methods=["POST"])
def cahier_charges_add():
    title = request.form.get("title", "").strip()
    description = request.form.get("description", "").strip()
    player_id = request.form.get("player_id") or None
    if not title:
        flash("Merci d'indiquer au moins un titre.", "error")
    else:
        db = get_db()
        db.execute(
            """INSERT INTO charges_items (title, description, status, created_by, created_at, player_id)
               VALUES (%s, %s, 'a_faire', %s, %s, %s)""",
            (title, description or None, session.get("user_email", ""), datetime.utcnow().isoformat(), player_id),
        )
        db.commit()
        flash("Tâche ajoutée.", "success")
    return redirect(url_for("cahier_charges", joueur=player_id))

@app.route("/cahier-des-charges/<int:item_id>/statut", methods=["POST"])
def cahier_charges_status(item_id):
    new_status = request.form.get("status", "")
    if new_status not in CHARGES_STATUSES:
        abort(400)
    db = get_db()
    row = db.execute("SELECT * FROM charges_items WHERE id = %s", (item_id,)).fetchone()
    if not row:
        abort(404)
    db.execute(
        "UPDATE charges_items SET status = %s, updated_at = %s WHERE id = %s",
        (new_status, datetime.utcnow().isoformat(), item_id),
    )
    db.commit()
    flash(f"Tâche déplacée vers « {CHARGES_STATUS_LABELS[new_status]} ».", "success")
    return redirect(url_for("cahier_charges", joueur=row["player_id"]))

@app.route("/cahier-des-charges/<int:item_id>/supprimer", methods=["POST"])
def cahier_charges_delete(item_id):
    db = get_db()
    row = db.execute("SELECT * FROM charges_items WHERE id = %s", (item_id,)).fetchone()
    if not row:
        abort(404)
    player_id = row["player_id"]
    if not _doc_can_delete(dict(row)):
        flash("Tu ne peux supprimer que les tâches que tu as créées (ou demande à l'admin).", "error")
    else:
        db.execute("DELETE FROM charges_items WHERE id = %s", (item_id,))
        db.commit()
        flash("Tâche supprimée.", "success")
    return redirect(url_for("cahier_charges", joueur=player_id))

@app.route("/cahier-des-charges/joueur/<int:player_id>/document", methods=["POST"])
def cahier_charges_upload(player_id):
    db = get_db()
    player = db.execute("SELECT * FROM players WHERE id = %s", (player_id,)).fetchone()
    if not player:
        abort(404)
    files = [f for f in request.files.getlist("files") if f and f.filename]
    if not files:
        flash("Merci de sélectionner au moins un fichier.", "error")
        return redirect(url_for("cahier_charges", joueur=player_id))
    saved = 0
    for file in files:
        payload = file.read()
        if not payload:
            continue
        db.execute(
            """INSERT INTO documents
               (kind, filename, mimetype, size_bytes, data, uploaded_by, uploaded_at,
                visibility, shared_player_id)
               VALUES ('file', %s, %s, %s, %s, %s, %s, 'player', %s)""",
            (
                file.filename, file.mimetype or "application/octet-stream",
                len(payload), psycopg2.Binary(payload),
                session.get("user_email", ""), datetime.utcnow().isoformat(), player_id,
            ),
        )
        saved += 1
    db.commit()
    flash(
        f"{saved} fichier{'s' if saved > 1 else ''} déposé{'s' if saved > 1 else ''} pour "
        f"{player['first_name']} {player['last_name']} — visible par lui dans « Mes documents ».",
        "success",
    )
    return redirect(url_for("cahier_charges", joueur=player_id))

@app.route("/cahier-des-charges/joueur/<int:player_id>/lien", methods=["POST"])
def cahier_charges_add_link(player_id):
    db = get_db()
    player = db.execute("SELECT * FROM players WHERE id = %s", (player_id,)).fetchone()
    if not player:
        abort(404)
    title = request.form.get("title", "").strip()
    url = request.form.get("url", "").strip()
    if not url or not (url.startswith("http://") or url.startswith("https://")):
        flash("Merci de coller un lien complet (commençant par http:// ou https://).", "error")
    else:
        db.execute(
            """INSERT INTO documents (kind, url, title, uploaded_by, uploaded_at, visibility, shared_player_id)
               VALUES ('link', %s, %s, %s, %s, 'player', %s)""",
            (url, title or url, session.get("user_email", ""), datetime.utcnow().isoformat(), player_id),
        )
        db.commit()
        flash(f"Lien ajouté — visible par {player['first_name']} dans « Mes documents ».", "success")
    return redirect(url_for("cahier_charges", joueur=player_id))

@app.route("/cahier-des-charges/document/<int:doc_id>/supprimer", methods=["POST"])
def cahier_charges_doc_delete(doc_id):
    db = get_db()
    doc = db.execute("SELECT * FROM documents WHERE id = %s", (doc_id,)).fetchone()
    if not doc:
        abort(404)
    player_id = doc["shared_player_id"]
    if not _doc_can_delete(dict(doc)):
        flash("Tu ne peux supprimer que tes propres dépôts (ou demande à l'admin).", "error")
    else:
        db.execute("DELETE FROM documents WHERE id = %s", (doc_id,))
        db.commit()
        flash("Document supprimé.", "success")
    return redirect(url_for("cahier_charges", joueur=player_id))

@app.route("/cahier-des-charges/joueur/<int:player_id>/evaluation-rugby/ajouter", methods=["POST"])
def ppid_rugby_add(player_id):
    db = get_db()
    player = db.execute("SELECT * FROM players WHERE id = %s", (player_id,)).fetchone()
    if not player:
        abort(404)
    period_label = request.form.get("period_label", "").strip()
    if not period_label:
        flash("Merci d'indiquer un intitulé de période (ex. « Novembre - Janvier »).", "error")
        return redirect(url_for("cahier_charges", joueur=player_id))
    eval_date = request.form.get("eval_date", "").strip() or None
    ratings = _ppid_rugby_ratings_from_form(request.form)
    objectifs = request.form.get("objectifs", "").strip() or None
    entrainements = request.form.get("entrainements", "").strip() or None
    db.execute(
        """INSERT INTO ppid_rugby_evals
           (player_id, period_label, eval_date, ratings, objectifs, entrainements, created_by, created_at)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
        (player_id, period_label, eval_date, ratings, objectifs, entrainements,
         session.get("user_email", ""), datetime.utcnow().isoformat()),
    )
    db.commit()
    flash(f"Point d'étape rugby ajouté pour {player['first_name']} {player['last_name']}.", "success")
    return redirect(url_for("cahier_charges", joueur=player_id))

@app.route("/cahier-des-charges/evaluation-rugby/<int:eval_id>/modifier", methods=["POST"])
def ppid_rugby_edit(eval_id):
    db = get_db()
    row = db.execute("SELECT * FROM ppid_rugby_evals WHERE id = %s", (eval_id,)).fetchone()
    if not row:
        abort(404)
    period_label = request.form.get("period_label", "").strip() or row["period_label"]
    eval_date = request.form.get("eval_date", "").strip() or None
    ratings = _ppid_rugby_ratings_from_form(request.form)
    objectifs = request.form.get("objectifs", "").strip() or None
    entrainements = request.form.get("entrainements", "").strip() or None
    db.execute(
        """UPDATE ppid_rugby_evals SET period_label = %s, eval_date = %s, ratings = %s,
           objectifs = %s, entrainements = %s, updated_at = %s WHERE id = %s""",
        (period_label, eval_date, ratings, objectifs, entrainements, datetime.utcnow().isoformat(), eval_id),
    )
    db.commit()
    flash("Point d'étape rugby mis à jour.", "success")
    return redirect(url_for("cahier_charges", joueur=row["player_id"]))

@app.route("/cahier-des-charges/evaluation-rugby/<int:eval_id>/supprimer", methods=["POST"])
def ppid_rugby_delete(eval_id):
    db = get_db()
    row = db.execute("SELECT * FROM ppid_rugby_evals WHERE id = %s", (eval_id,)).fetchone()
    if not row:
        abort(404)
    db.execute("DELETE FROM ppid_rugby_evals WHERE id = %s", (eval_id,))
    db.commit()
    flash("Point d'étape rugby supprimé.", "success")
    return redirect(url_for("cahier_charges", joueur=row["player_id"]))

@app.route("/cahier-des-charges/joueur/<int:player_id>/evaluation-physique/ajouter", methods=["POST"])
def ppid_physical_add(player_id):
    db = get_db()
    player = db.execute("SELECT * FROM players WHERE id = %s", (player_id,)).fetchone()
    if not player:
        abort(404)
    period_label = request.form.get("period_label", "").strip()
    if not period_label:
        flash("Merci d'indiquer un intitulé de période (ex. « Novembre - Janvier »).", "error")
        return redirect(url_for("cahier_charges", joueur=player_id))
    eval_date = request.form.get("eval_date", "").strip() or None
    ratings = _ppid_physical_coach_ratings_from_form(request.form, "{}")
    commentaires = request.form.get("commentaires", "").strip() or None
    axe_musculation = request.form.get("axe_musculation", "").strip() or None
    axe_terrain = request.form.get("axe_terrain", "").strip() or None
    db.execute(
        """INSERT INTO ppid_physical_evals
           (player_id, period_label, eval_date, ratings, commentaires, axe_musculation, axe_terrain,
            created_by, created_at)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
        (player_id, period_label, eval_date, ratings, commentaires, axe_musculation, axe_terrain,
         session.get("user_email", ""), datetime.utcnow().isoformat()),
    )
    db.commit()
    flash(f"Point d'étape physique ajouté pour {player['first_name']} {player['last_name']}.", "success")
    return redirect(url_for("cahier_charges", joueur=player_id))

@app.route("/cahier-des-charges/evaluation-physique/<int:eval_id>/modifier", methods=["POST"])
def ppid_physical_edit(eval_id):
    db = get_db()
    row = db.execute("SELECT * FROM ppid_physical_evals WHERE id = %s", (eval_id,)).fetchone()
    if not row:
        abort(404)
    period_label = request.form.get("period_label", "").strip() or row["period_label"]
    eval_date = request.form.get("eval_date", "").strip() or None
    ratings = _ppid_physical_coach_ratings_from_form(request.form, row["ratings"])
    commentaires = request.form.get("commentaires", "").strip() or None
    axe_musculation = request.form.get("axe_musculation", "").strip() or None
    axe_terrain = request.form.get("axe_terrain", "").strip() or None
    db.execute(
        """UPDATE ppid_physical_evals SET period_label = %s, eval_date = %s, ratings = %s,
           commentaires = %s, axe_musculation = %s, axe_terrain = %s, updated_at = %s WHERE id = %s""",
        (period_label, eval_date, ratings, commentaires, axe_musculation, axe_terrain,
         datetime.utcnow().isoformat(), eval_id),
    )
    db.commit()
    flash("Point d'étape physique mis à jour.", "success")
    return redirect(url_for("cahier_charges", joueur=row["player_id"]))

@app.route("/cahier-des-charges/evaluation-physique/<int:eval_id>/supprimer", methods=["POST"])
def ppid_physical_delete(eval_id):
    db = get_db()
    row = db.execute("SELECT * FROM ppid_physical_evals WHERE id = %s", (eval_id,)).fetchone()
    if not row:
        abort(404)
    db.execute("DELETE FROM ppid_physical_evals WHERE id = %s", (eval_id,))
    db.commit()
    flash("Point d'étape physique supprimé.", "success")
    return redirect(url_for("cahier_charges", joueur=row["player_id"]))

@app.route("/cahier-des-charges/joueur/<int:player_id>/entretien/ajouter", methods=["POST"])
def ppid_entretien_add(player_id):
    db = get_db()
    player = db.execute("SELECT * FROM players WHERE id = %s", (player_id,)).fetchone()
    if not player:
        abort(404)
    entretien_date = request.form.get("entretien_date", "").strip()
    entretien_type = request.form.get("entretien_type", "").strip()
    notes = request.form.get("notes", "").strip()
    if not entretien_date or entretien_type not in PPID_ENTRETIEN_TYPES:
        flash("Merci d'indiquer une date et un type d'entretien valides.", "error")
        return redirect(url_for("cahier_charges", joueur=player_id))
    db.execute(
        """INSERT INTO ppid_entretiens (player_id, entretien_date, entretien_type, notes, created_by, created_at)
           VALUES (%s, %s, %s, %s, %s, %s)""",
        (player_id, entretien_date, entretien_type, notes or None, session.get("user_email", ""), datetime.utcnow().isoformat()),
    )
    db.commit()
    flash(f"Entretien ajouté au cahier de {player['first_name']} — visible par lui dans « Mes évaluations ».", "success")
    return redirect(url_for("cahier_charges", joueur=player_id))

@app.route("/cahier-des-charges/entretien/<int:entretien_id>/modifier", methods=["POST"])
def ppid_entretien_edit(entretien_id):
    db = get_db()
    row = db.execute("SELECT * FROM ppid_entretiens WHERE id = %s", (entretien_id,)).fetchone()
    if not row:
        abort(404)
    entretien_date = request.form.get("entretien_date", "").strip() or row["entretien_date"]
    entretien_type = request.form.get("entretien_type", "").strip()
    if entretien_type not in PPID_ENTRETIEN_TYPES:
        entretien_type = row["entretien_type"]
    notes = request.form.get("notes", "").strip()
    db.execute(
        "UPDATE ppid_entretiens SET entretien_date = %s, entretien_type = %s, notes = %s, updated_at = %s WHERE id = %s",
        (entretien_date, entretien_type, notes or None, datetime.utcnow().isoformat(), entretien_id),
    )
    db.commit()
    flash("Entretien mis à jour.", "success")
    return redirect(url_for("cahier_charges", joueur=row["player_id"]))

@app.route("/cahier-des-charges/entretien/<int:entretien_id>/supprimer", methods=["POST"])
def ppid_entretien_delete(entretien_id):
    db = get_db()
    row = db.execute("SELECT * FROM ppid_entretiens WHERE id = %s", (entretien_id,)).fetchone()
    if not row:
        abort(404)
    db.execute("DELETE FROM ppid_entretiens WHERE id = %s", (entretien_id,))
    db.commit()
    flash("Entretien supprimé.", "success")
    return redirect(url_for("cahier_charges", joueur=row["player_id"]))

# ---------------------------------------------------------------------------
# GESTION DES JOUEURS (ADMIN) — effectif (ajout manuel + import Excel), groupes
# personnalisables (Avants/Trois-quarts par défaut, l'admin peut en créer
# d'autres), et actions sur un joueur (changer de groupe, réinitialiser le mot
# de passe, supprimer le compte). Réservé à l'admin (@admin_required) : le
# staff est bien connecté mais ne gère pas les comptes joueurs.
# ---------------------------------------------------------------------------
# Postes de terrain -> catégorie (mêmes 7 catégories que SQUAD_ROSTER dans
# parser.py). Ordre de test volontaire pour lever les ambiguïtés des postes
# polyvalents saisis dans le fichier Excel du club (ex. "Pilier - Talonneur"
# doit ressortir en Talonneur, "Pilier droit / 2ème ligne" doit ressortir en
# Pilier) : on retient la 1ère catégorie qui correspond, dans cet ordre.
def _classify_player_position(poste):
    p = (poste or "").lower()
    if "talonneur" in p:
        return "Talonneur"
    if "mêlée" in p or "melee" in p or "ouverture" in p:
        return "Charnière"
    if "pilier" in p:
        return "Pilier"
    has_l = bool(re.search(r"\bl\b", p)) or "ligne" in p
    if "2" in p and has_l:
        return "2ème ligne"
    if "3" in p and has_l:
        return "3ème ligne"
    if "ailier" in p or "arrière" in p or "arriere" in p:
        return "Ailier/Arrière"
    if "centre" in p:
        return "Centre"
    return None

# Classification plus fine (10 postes du P.P.I.D — voir PPID_POSITIONS) que
# _classify_player_position ci-dessus (7 catégories, seulement pour Avants/Trois-quarts) :
# réutilise le même texte « Poste » du fichier Excel du club pour pré-remplir le poste PPID
# à l'import, plutôt que de faire régler les ~60 joueurs un par un à la main. Peut renvoyer
# None quand le texte ne permet pas de trancher (ex. « 3ème ligne » seul, sans « aile » ni
# « centre », ou « Charnière » seul, sans « mêlée » ni « ouverture ») : le poste reste alors
# à régler manuellement pour ce joueur (voir le message de fin d'import).
def _classify_ppid_position(poste):
    p = (poste or "").lower()
    has_melee = "mêlée" in p or "melee" in p
    has_ouverture = "ouverture" in p
    if has_melee and not has_ouverture:
        return "9"
    if has_ouverture and not has_melee:
        return "10"
    if "talonneur" in p:
        return "Talon"
    if "pilier" in p:
        return "Pilier"
    has_l = bool(re.search(r"\bl\b", p)) or "ligne" in p
    if "2" in p and has_l:
        return "2L"
    if "3" in p and has_l:
        if "centre" in p:
            return "3LC"
        if "aile" in p:
            return "3LA"
        return None
    has_ailier = "ailier" in p or "aile" in p
    has_arriere = "arrière" in p or "arriere" in p
    if has_ailier and not has_arriere:
        return "11-14"
    if has_arriere and not has_ailier:
        return "15"
    if "centre" in p:
        return "12-13"
    return None

# Regroupement par défaut avant/trois-quarts (convention rugby classique), utilisé
# uniquement pour proposer un groupe de partage de documents à l'import — l'admin
# peut ensuite réaffecter n'importe quel joueur à n'importe quel groupe à la main.
_FORWARD_CATEGORIES = {"Pilier", "Talonneur", "2ème ligne", "3ème ligne"}

def _default_group_name_for_category(cat):
    if cat is None:
        return None
    return "Avants" if cat in _FORWARD_CATEGORIES else "Trois-quarts"

def _split_player_name(full_name):
    """Sépare 'NOM Prénom' (format du fichier Excel du club, nom de famille en
    MAJUSCULES en tête) en (prénom, nom). Gère les noms de famille à plusieurs mots
    (ex. 'PERDIGON LE NAOUR Oscar') et les prénoms composés (ex. 'FUKWAMOKO Mauricio
    Lorenzo') : le nom de famille est la suite de mots ENTIÈREMENT EN MAJUSCULES en
    début de chaîne, le reste est le prénom. Si toute la chaîne est en majuscules
    (mauvaise saisie), le dernier mot est pris comme prénom par défaut."""
    tokens = (full_name or "").split()
    if not tokens:
        return "", ""
    idx = 0
    while idx < len(tokens) and tokens[idx].isupper():
        idx += 1
    if idx == 0:
        idx = 1
    if idx >= len(tokens):
        idx = max(1, len(tokens) - 1)
    last = " ".join(tokens[:idx]).title()
    first = " ".join(tokens[idx:])
    return first, last

@app.route("/admin/joueurs")
@admin_required
def admin_joueurs():
    db = get_db()
    groups = db.execute("SELECT * FROM player_groups ORDER BY name").fetchall()
    players = db.execute(
        """SELECT p.*, g.name AS group_name FROM players p
           LEFT JOIN player_groups g ON g.id = p.group_id
           ORDER BY p.last_name, p.first_name"""
    ).fetchall()
    return render_template("admin_joueurs.html", groups=groups, players=players, ppid_positions=PPID_POSITIONS)

@app.route("/admin/joueurs/ajouter", methods=["POST"])
@admin_required
def admin_joueurs_ajouter():
    first_name = request.form.get("first_name", "").strip()
    last_name = request.form.get("last_name", "").strip()
    email = request.form.get("email", "").strip().lower()
    group_id = request.form.get("group_id") or None
    if not first_name or not last_name or not email or "@" not in email:
        flash("Merci de renseigner prénom, nom et un email valide.", "error")
        return redirect(url_for("admin_joueurs"))
    db = get_db()
    existing = db.execute("SELECT id FROM players WHERE email = %s", (email,)).fetchone()
    if existing:
        flash("Un joueur avec cet email existe déjà.", "error")
        return redirect(url_for("admin_joueurs"))
    db.execute(
        "INSERT INTO players (first_name, last_name, email, group_id, created_at) VALUES (%s, %s, %s, %s, %s)",
        (first_name, last_name, email, group_id, datetime.utcnow().isoformat()),
    )
    db.commit()
    flash(f"{first_name} {last_name} ajouté. Il pourra se connecter avec {email} et choisira son mot de passe à la 1ère connexion.", "success")
    return redirect(url_for("admin_joueurs"))

@app.route("/admin/joueurs/importer", methods=["POST"])
@admin_required
def admin_joueurs_importer():
    file = request.files.get("fichier")
    if not file or not file.filename:
        flash("Merci de sélectionner un fichier Excel (.xlsx).", "error")
        return redirect(url_for("admin_joueurs"))
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.worksheets[0]
    except Exception:
        flash("Fichier Excel illisible. Vérifie que c'est bien un fichier .xlsx.", "error")
        return redirect(url_for("admin_joueurs"))
    # Repère la ligne d'en-têtes (contient "Mail" ou "Email") pour situer les colonnes
    # NOM Prénom / Poste / Mail, plutôt que de figer des numéros de colonnes qui
    # casseraient si le fichier du club change légèrement de mise en page.
    header_row, col_nom, col_poste, col_mail = None, None, None, None
    for r in range(1, min(ws.max_row, 10) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        for c, v in enumerate(row_vals, start=1):
            if not isinstance(v, str):
                continue
            low = v.strip().lower()
            if "mail" in low:
                col_mail = c
            elif "poste" in low:
                col_poste = c
            elif "nom" in low and col_nom is None:
                col_nom = c
        if col_mail:
            header_row = r
            break
    if not header_row or not col_nom or not col_mail:
        flash("Colonnes attendues introuvables (il faut au moins une colonne « NOM Prénom » et une colonne « Mail »).", "error")
        return redirect(url_for("admin_joueurs"))
    db = get_db()
    groups = {g["name"]: g["id"] for g in db.execute("SELECT id, name FROM player_groups").fetchall()}
    created, updated, uncategorized, ppid_unresolved = 0, 0, [], []
    for r in range(header_row + 1, ws.max_row + 1):
        nom_prenom = ws.cell(row=r, column=col_nom).value
        mail = ws.cell(row=r, column=col_mail).value
        poste = ws.cell(row=r, column=col_poste).value if col_poste else None
        if not nom_prenom or not mail or "@" not in str(mail):
            continue
        email = str(mail).strip().lower()
        first, last = _split_player_name(str(nom_prenom).strip())
        if not last:
            continue
        cat = _classify_player_position(str(poste or ""))
        if cat is None:
            uncategorized.append(f"{first} {last}".strip())
        group_name = _default_group_name_for_category(cat)
        group_id = groups.get(group_name) if group_name else None
        # Poste PPID (plus fin que le groupe Avants/Trois-quarts) déduit du même texte
        # « Poste » du fichier Excel — voir _classify_ppid_position. Repéré séparément de
        # `cat` ci-dessus : un joueur peut être correctement classé Avant/Trois-quarts
        # (cat non None) tout en restant ambigu pour le poste PPID précis (ex. « 3ème
        # ligne » sans préciser aile/centre), et inversement.
        ppid_pos = _classify_ppid_position(str(poste or ""))
        if ppid_pos is None:
            ppid_unresolved.append(f"{first} {last}".strip())
        existing = db.execute("SELECT id, group_id FROM players WHERE email = %s", (email,)).fetchone()
        if existing:
            # COALESCE(ppid_position, ...) : ne remplit que si le poste n'a encore jamais
            # été réglé, pour ne jamais écraser une correction manuelle faite par l'admin
            # depuis la page Gestion des joueurs à un import précédent.
            db.execute(
                "UPDATE players SET first_name = %s, last_name = %s, ppid_position = COALESCE(ppid_position, %s) WHERE id = %s",
                (first, last, ppid_pos, existing["id"]),
            )
            updated += 1
        else:
            db.execute(
                """INSERT INTO players (first_name, last_name, email, group_id, ppid_position, created_at)
                   VALUES (%s, %s, %s, %s, %s, %s)""",
                (first, last, email, group_id, ppid_pos, datetime.utcnow().isoformat()),
            )
            created += 1
    db.commit()
    msg = f"Import terminé : {created} joueur{'s' if created != 1 else ''} ajouté{'s' if created != 1 else ''}, {updated} mis à jour."
    if uncategorized:
        msg += f" ⚠️ Poste non reconnu (groupe non assigné automatiquement) pour : {', '.join(uncategorized)}."
    if ppid_unresolved:
        msg += (
            f" ⚠️ Poste PPID pas assez précis pour distinguer automatiquement (ex. « 3ème ligne » sans aile/centre, "
            f"« Charnière » sans mêlée/ouverture) — à régler à la main dans le tableau ci-dessous pour : "
            f"{', '.join(ppid_unresolved)}."
        )
    flash(msg, "success")
    return redirect(url_for("admin_joueurs"))

@app.route("/admin/joueurs/<int:player_id>/groupe", methods=["POST"])
@admin_required
def admin_joueurs_groupe(player_id):
    db = get_db()
    player = db.execute("SELECT * FROM players WHERE id = %s", (player_id,)).fetchone()
    if not player:
        abort(404)
    group_id = request.form.get("group_id") or None
    db.execute("UPDATE players SET group_id = %s WHERE id = %s", (group_id, player_id))
    db.commit()
    flash(f"Groupe mis à jour pour {player['first_name']} {player['last_name']}.", "success")
    return redirect(url_for("admin_joueurs"))

@app.route("/admin/joueurs/<int:player_id>/poste-ppid", methods=["POST"])
@admin_required
def admin_joueurs_poste_ppid(player_id):
    db = get_db()
    player = db.execute("SELECT * FROM players WHERE id = %s", (player_id,)).fetchone()
    if not player:
        abort(404)
    poste = request.form.get("ppid_position") or None
    if poste and poste not in PPID_POSITIONS:
        abort(400)
    db.execute("UPDATE players SET ppid_position = %s WHERE id = %s", (poste, player_id))
    db.commit()
    flash(f"Poste PPID mis à jour pour {player['first_name']} {player['last_name']}.", "success")
    return redirect(url_for("admin_joueurs"))

@app.route("/admin/joueurs/<int:player_id>/reinitialiser-mdp", methods=["POST"])
@admin_required
def admin_joueurs_reset_password(player_id):
    db = get_db()
    player = db.execute("SELECT * FROM players WHERE id = %s", (player_id,)).fetchone()
    if not player:
        abort(404)
    db.execute("UPDATE players SET password_hash = NULL WHERE id = %s", (player_id,))
    db.commit()
    flash(f"Mot de passe réinitialisé pour {player['first_name']} {player['last_name']} : il en choisira un nouveau à sa prochaine connexion.", "success")
    return redirect(url_for("admin_joueurs"))

@app.route("/admin/joueurs/<int:player_id>/supprimer", methods=["POST"])
@admin_required
def admin_joueurs_supprimer(player_id):
    db = get_db()
    player = db.execute("SELECT * FROM players WHERE id = %s", (player_id,)).fetchone()
    if not player:
        abort(404)
    db.execute("DELETE FROM players WHERE id = %s", (player_id,))
    db.commit()
    flash(f"Compte de {player['first_name']} {player['last_name']} supprimé.", "success")
    return redirect(url_for("admin_joueurs"))

@app.route("/admin/groupes/creer", methods=["POST"])
@admin_required
def admin_groupes_creer():
    name = request.form.get("name", "").strip()
    if not name:
        flash("Merci d'indiquer un nom de groupe.", "error")
        return redirect(url_for("admin_joueurs"))
    db = get_db()
    existing = db.execute("SELECT id FROM player_groups WHERE name = %s", (name,)).fetchone()
    if existing:
        flash(f"Le groupe « {name} » existe déjà.", "error")
        return redirect(url_for("admin_joueurs"))
    db.execute(
        "INSERT INTO player_groups (name, created_at) VALUES (%s, %s)",
        (name, datetime.utcnow().isoformat()),
    )
    db.commit()
    flash(f"Groupe « {name} » créé.", "success")
    return redirect(url_for("admin_joueurs"))

@app.route("/admin/groupes/<int:group_id>/supprimer", methods=["POST"])
@admin_required
def admin_groupes_supprimer(group_id):
    db = get_db()
    group = db.execute("SELECT * FROM player_groups WHERE id = %s", (group_id,)).fetchone()
    if not group:
        abort(404)
    db.execute("DELETE FROM player_groups WHERE id = %s", (group_id,))
    db.commit()
    flash(f"Groupe « {group['name']} » supprimé. Les joueurs de ce groupe n'ont plus de groupe assigné.", "success")
    return redirect(url_for("admin_joueurs"))

init_db()
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
