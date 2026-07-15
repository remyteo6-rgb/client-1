"""Parseur du rapport hebdomadaire Pro D2 (fichier Excel envoyé par la ligue/le prestataire
data). Ce fichier est totalement indépendant de parser.py (qui lit les exports Sportscode
XML de nos propres matchs) : ici on lit un classeur Excel qui contient, pour toute la
Pro D2 :
  - un onglet "Classement" (une ligne par équipe : points, V/N/D, essais, cartons...)
  - une vingtaine d'onglets de comparaison où chaque ligne est une équipe (Attaque,
    Défense, Discipline, Touches, Mêlées, Rucks, Jeu au pied...), avec en plus une ligne
    "ProD2" qui donne la moyenne de la ligue sur cet onglet ;
  - un onglet par équipe (nommé exactement comme l'équipe) qui liste tous les joueurs de
    son effectif avec leurs statistiques individuelles.

Le parseur ne code pas la liste des 16 équipes en dur : il reconnaît le type de chaque
onglet à son premier en-tête ("Equipes" = onglet de comparaison, "Nom du joueur" = onglet
joueurs d'une équipe). Ça permet au rapport de continuer à fonctionner d'une semaine à
l'autre même si une équipe change de nom, ou en cas de montée/descente d'une saison à
l'autre.
"""

import unicodedata

import openpyxl


def _normalize(name):
    """Compare deux noms d'équipe en ignorant accents/casse/espaces. Utile car le
    classeur Excel n'est pas toujours cohérent d'un onglet à l'autre : par exemple
    l'onglet Classement écrit "AS Béziers Hérault" (avec accents) alors que l'onglet
    joueurs de cette équipe et tous les onglets de comparaison l'écrivent "AS Beziers
    Herault" (sans accents). Sans cette normalisation, la fiche de cette équipe se
    retrouverait avec un classement manquant."""
    if not name:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(name))
    ascii_name = "".join(c for c in nfkd if not unicodedata.combining(c))
    return " ".join(ascii_name.strip().lower().split())


def parse_prod2_report(file_path):
    """Lit le classeur Excel et renvoie un dict :
    {
      "classement": [ {colonne: valeur, ...}, ... ]  # une ligne par équipe, dans l'ordre du fichier
      "category_sheets": { "Attaque": {"CA Brive": {col: val, ...}, "ProD2": {...}, ...}, ... }
      "teams_players": { "CA Brive": [ {col: val, ...}, ... ], ... }
      "team_names": ["CA Brive", "AS Beziers Herault", ...]  # dans l'ordre du classement
    }
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    result = {"classement": [], "category_sheets": {}, "teams_players": {}, "team_names": []}
    canonical_by_normalized = {}

    if "Classement" in wb.sheetnames:
        ws = wb["Classement"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            result["classement"].append(dict(zip(headers, row)))
        result["team_names"] = [r.get("Nom de l'équipe") for r in result["classement"] if r.get("Nom de l'équipe")]
        canonical_by_normalized = {_normalize(n): n for n in result["team_names"]}

    def _canonical(team_key):
        """Fait correspondre un nom d'équipe rencontré dans un autre onglet à son nom
        exact tel qu'écrit dans le Classement (source de référence pour l'affichage).
        'ProD2' (ligne de moyenne de ligue) n'est jamais une équipe : renvoyé tel quel."""
        if team_key == "ProD2":
            return team_key
        return canonical_by_normalized.get(_normalize(team_key), team_key)

    for name in wb.sheetnames:
        if name == "Classement":
            continue
        ws = wb[name]
        first_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if not first_row:
            continue
        headers = [c.value for c in first_row]
        if not headers or headers[0] is None:
            continue
        first_header = headers[0]

        if first_header == "Equipes":
            rows = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                team = _canonical(row[0])
                rows[team] = dict(zip(headers[1:], row[1:]))
            result["category_sheets"][name] = rows
        elif first_header == "Nom du joueur":
            players = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                players.append(dict(zip(headers, row)))
            team_name = _canonical(name)
            result["teams_players"][team_name] = players
            if team_name not in result["team_names"]:
                result["team_names"].append(team_name)

    return result


def _num(v):
    """Arrondit proprement les nombres pour l'affichage (les moyennes de ligue arrivent
    avec plein de décimales, ex: 37.0625) ; laisse tel quel le texte (ex: '36min44')."""
    if isinstance(v, float):
        if v == int(v):
            return int(v)
        return round(v, 1)
    return v


def get_team_profile(report, team_name):
    """Regroupe, pour UNE équipe, sa ligne de classement + sa ligne dans chaque onglet de
    comparaison (avec la moyenne ProD2 en vis-à-vis) + la liste de ses joueurs."""
    profile = {
        "name": team_name,
        "classement": None,
        "categories": {},   # {sheet_name: {col: value, ...}}
        "league_avg": {},   # {sheet_name: {col: value, ...}}
        "players": [],
    }

    for row in report["classement"]:
        if row.get("Nom de l'équipe") == team_name:
            profile["classement"] = {k: _num(v) for k, v in row.items()}
            break

    for sheet_name, rows in report["category_sheets"].items():
        if team_name in rows:
            profile["categories"][sheet_name] = {k: _num(v) for k, v in rows[team_name].items()}
        if "ProD2" in rows:
            profile["league_avg"][sheet_name] = {k: _num(v) for k, v in rows["ProD2"].items()}

    profile["players"] = [
        {k: _num(v) for k, v in p.items()} for p in report["teams_players"].get(team_name, [])
    ]
    return profile


def get_classement_table(report):
    """Classement complet (déjà dans l'ordre du fichier), valeurs arrondies pour l'affichage."""
    return [{k: _num(v) for k, v in row.items()} for row in report["classement"]]


# ---- Regroupement des onglets bruts par page du site --------------------------------
# Chaque page "secteur" d'une fiche équipe affiche 1 ou plusieurs de ces tableaux bruts
# les uns en dessous des autres (comparés à la moyenne ProD2). On ne renomme aucune
# colonne : on affiche exactement les intitulés du fichier Excel, pour rester fidèle au
# rapport source et ne rien avoir à retraduire à la main chaque semaine.
SECTOR_SHEETS = {
    "attaque": ["Attaque", "Essais Marqués", "Points Marqués"],
    "defense": ["Défense", "Essais Encaissés", "Points Encaissés"],
    "discipline": ["Discipline"],
    "touches": ["Touches", "Touches - Alignements", "Touches - lancements", "Touches défensives"],
    "melee": ["Mêlées", "Mêlées - lancements", "Mêlées défensives"],
    "rucks": ["Rucks", "Rucks - Joueurs Consommés", "Rucks - Vitesse de Libération", "Rucks défensifs"],
    "jap": ["Jeux au pied", "Duels aériens", "Penaltouches", "Tirs au but", "restarts", "restarts_defense"],
}

# Colonnes des onglets joueurs, regroupées par thème pour l'affichage (mêmes intitulés
# que le fichier Excel). Toute colonne du fichier qui n'apparaît dans aucun groupe est
# ajoutée automatiquement à "Général" en fin de liste (voir compute_player_groups).
PLAYER_COLUMN_GROUPS = {
    "Général": ["Matchs", "Temps de Jeu", "Titulaires", "Remplaçants", "Points Marqués", "Essais"],
    "Attaque": [
        "Passes", "Passes positives", "Offloads", "Contacts", "Contacts positifs",
        "Contacts réussis %", "Soutiens offensifs", "Soutiens offensifs réussis",
        "Soutiens offensifs réussis %", "Défenseurs battus", "Franchissements", "Mètres parcourus",
        "Ballons perdus", "En avant",
    ],
    "Défense": [
        "Plaquages", "Premiers plaquages", "Plaquages assistants", "Plaquages réussis",
        "Premiers plaquages réussis", "Plaquages assistants réussis", "Plaquages réussis %",
        "Premiers plaquages réussis %", "Plaquages assistants réussis %", "Plaquages au sol",
        "Plaquages actifs", "Plaquages actifs %", "Plaquages avancés", "Plaquages subis",
    ],
    "Ruck / Contest": [
        "Rucks", "Rucks Positifs", "Rucks réussis %", "Contests", "Contests positifs %",
        "Contests positifs", "Contests négatifs", "Contre-rucks", "Contre-rucks positifs %",
        "Contre-rucks positifs", "Contre-rucks négatifs",
    ],
    "Touche / Mêlée": [
        "Lancers", "Lancers réussis", "Lancers réussis %", "Ballons pris en touche",
        "Ballons volés en touche", "Mêlées offensives positives", "Mêlées offensives",
        "Mêlées défensives positives", "Mêlées défensives",
    ],
    "Jeu au pied": [
        "Jeu au pied", "Distance jeux au pied", "50/22", "Duels aériens",
        "Duels aériens gagnés %", "Duels aériens offensifs", "Duels aériens offensifs gagnés %",
        "Duels aériens défensifs", "Duels aériens défensifs gagnés %", "Tirs au but",
        "Tirs au but réussis", "Tirs au but réussis %", "Drops",
    ],
    "Discipline": ["Pénalités concédées", "Cartons jaunes", "Cartons oranges", "Cartons rouges"],
}
# ---- Tendances Pro D2 (points forts / points faibles par équipe) --------------------
# Sélection volontairement restreinte à des indicateurs qui ont un sens clair ("mieux vaut
# plus" ou "mieux vaut moins"), plutôt que les ~300 colonnes brutes du fichier (beaucoup
# n'ont pas de sens évident en soi, ex: nombre de rucks en zone 22m adverse). Format de
# chaque entrée : (onglet source, colonne, libellé affiché, higher_is_better).
PROD2_KPI_DEFINITIONS = [
    ("Classement", "Points", "Points au classement", True),
    ("Classement", "Essais Marqués", "Essais marqués (total)", True),
    ("Classement", "Essais Encaissés", "Essais encaissés (total)", False),
    ("Classement", "Cartons jaunes", "Cartons jaunes", False),
    ("Classement", "Cartons Rouges", "Cartons rouges", False),
    ("Possession", "% Possession", "Possession (%)", True),
    ("Possession", "% Occupation", "Occupation du terrain (%)", True),
    ("Attaque", "Défenseurs battus", "Défenseurs battus", True),
    ("Attaque", "Mètres parcourus", "Mètres parcourus", True),
    ("Attaque", "Offloads", "Offloads", True),
    ("Attaque", "Ballons perdus", "Ballons perdus", False),
    ("Attaque", "En avants", "En-avants", False),
    ("Attaque", "Points marqués/entrées dans les 22", "Points par entrée en 22m adverse", True),
    ("Essais Marqués", "Essais Marqués / Match", "Essais marqués par match", True),
    ("Défense", "Plaquages réussis %", "Plaquages réussis (%)", True),
    ("Défense", "Turnovers", "Turnovers gagnés", True),
    ("Défense", "Contests positifs %", "Contests aériens gagnés (%)", True),
    ("Défense", "Contre-rucks positifs %", "Contre-rucks gagnés (%)", True),
    ("Discipline", "Pénalités", "Pénalités concédées", False),
    ("Touches", "Touches gagnées %", "Touches gagnées (%)", True),
    ("Mêlées", "Mêlées gagnées %", "Mêlées gagnées (%)", True),
    ("Rucks", "Rucks gagnés %", "Rucks gagnés (%)", True),
    ("Rucks - Vitesse de Libération", "0-3s %", "Rucks rapides -3s (%)", True),
    ("Duels aériens", "Duels aériens gagnés %", "Duels aériens gagnés (%)", True),
    ("Tirs au but", "Tirs au but %", "Réussite au tir au but (%)", True),
    ("restarts", "restarts_won_perc", "Renvois récupérés (%)", True),
]


def _kpi_value(report, sheet, column, team_name):
    if sheet == "Classement":
        for row in report["classement"]:
            if row.get("Nom de l'équipe") == team_name:
                return row.get(column)
        return None
    rows = report["category_sheets"].get(sheet, {})
    row = rows.get(team_name)
    if row is None:
        return None
    return row.get(column)


# Dict {(onglet, colonne): higher_is_better} construit à partir de PROD2_KPI_DEFINITIONS,
# pour savoir en O(1), pour une ligne donnée d'un tableau de comparaison, si c'est un
# indicateur clé qu'on peut colorer en vert/rouge (et dans quel sens).
PROD2_KPI_DIRECTION = {
    (sheet, column): higher_is_better
    for sheet, column, _label, higher_is_better in PROD2_KPI_DEFINITIONS
}


def build_compare_rows(sheet_name, team_row, avg_row):
    """Transforme un tableau brut équipe/moyenne ProD2 (dict {colonne: valeur}) en liste de
    lignes prêtes à afficher, avec une couleur vert/rouge quand la ligne correspond à un
    indicateur clé (PROD2_KPI_DEFINITIONS) et qu'on peut la comparer à la moyenne de ligue.
    Les lignes qui ne sont pas des indicateurs clés, ou sans moyenne comparable, restent
    neutres (color=None) — on ne colore que ce qui a un sens de lecture clair."""
    rows = []
    for key, value in team_row.items():
        avg = avg_row.get(key)
        color = None
        higher_is_better = PROD2_KPI_DIRECTION.get((sheet_name, key))
        if (
            higher_is_better is not None
            and isinstance(value, (int, float))
            and isinstance(avg, (int, float))
            and value != avg
        ):
            better = value > avg if higher_is_better else value < avg
            color = "good" if better else "bad"
        rows.append({"key": key, "value": value, "avg": avg, "color": color})
    return rows


def compute_team_trends(report):
    """Pour chaque équipe, ressort les indicateurs (parmi PROD2_KPI_DEFINITIONS) où elle
    fait partie des 2 meilleures ou des 2 moins bonnes équipes de Pro D2 sur cet
    indicateur (sur 16 équipes), en respectant le sens de chaque indicateur (parfois plus
    c'est mieux, parfois moins c'est mieux). Un indicateur qui n'a pas assez de données
    cette semaine (moins de 8 équipes renseignées) est ignoré plutôt que faussé."""
    team_names = report["team_names"]
    trends = {t: {"strengths": [], "weaknesses": []} for t in team_names}

    for sheet, column, label, higher_is_better in PROD2_KPI_DEFINITIONS:
        values = {}
        for t in team_names:
            v = _kpi_value(report, sheet, column, t)
            if isinstance(v, (int, float)):
                values[t] = v
        if len(values) < 8:
            continue
        ranked = sorted(values.items(), key=lambda kv: kv[1], reverse=higher_is_better)
        n_ranked = len(ranked)
        for rank, (team_name, value) in enumerate(ranked, start=1):
            percentile = (n_ranked - rank) / (n_ranked - 1) * 100 if n_ranked > 1 else 50
            entry = {"label": label, "value": _num(value), "rank": rank, "n_teams": n_ranked}
            if percentile >= 90:
                trends[team_name]["strengths"].append(entry)
            elif percentile <= 10:
                trends[team_name]["weaknesses"].append(entry)
    return trends


def compute_team_kpi_profile(report, team_name):
    """Pour la vue d'ensemble d'une équipe : son percentile (0-100) sur chaque indicateur
    clé où on a assez de données, dans l'ordre de PROD2_KPI_DEFINITIONS. 50 = dans la
    moyenne de la Pro D2, 100 = la meilleure équipe sur cet indicateur, 0 = la moins bonne —
    toujours dans le bon sens, même pour les indicateurs où "moins c'est mieux" (ex:
    ballons perdus). Sert à afficher un graphique de profil sur la fiche équipe."""
    team_names = report["team_names"]
    labels, percentiles = [], []
    for sheet, column, label, higher_is_better in PROD2_KPI_DEFINITIONS:
        values = {}
        for t in team_names:
            v = _kpi_value(report, sheet, column, t)
            if isinstance(v, (int, float)):
                values[t] = v
        if team_name not in values or len(values) < 8:
            continue
        ranked = sorted(values.items(), key=lambda kv: kv[1], reverse=higher_is_better)
        n_ranked = len(ranked)
        rank = next(i for i, (t, _v) in enumerate(ranked, start=1) if t == team_name)
        percentile = round((n_ranked - rank) / (n_ranked - 1) * 100) if n_ranked > 1 else 50
        labels.append(label)
        percentiles.append(percentile)
    return {"labels": labels, "percentiles": percentiles}


# Colonnes utilisées pour repérer les joueurs à surveiller (menace tactique), différent de
# "les plus utilisés" (compute_most_used_players) qui ne regarde que le temps de jeu / les
# titularisations. Ici on cherche qui pèse le plus sur chaque registre du jeu.
PLAYER_THREAT_COLUMNS = [
    ("Essais", "Meilleur marqueur d'essais", "essais"),
    ("Mètres parcourus", "Meilleur porteur de balle", "m parcourus"),
    ("Défenseurs battus", "Meilleur perceur de ligne", "déf. battus"),
    ("Plaquages réussis", "Meilleur défenseur", "plaquages réussis"),
    ("Tirs au but réussis", "Buteur clé", "tirs réussis"),
]


def compute_team_threats(players):
    """Pour quelques colonnes clés du fichier joueurs, repère le joueur en tête de
    l'effectif — pas "qui joue le plus" (déjà couvert par compute_most_used_players) mais
    "qui pèse le plus" sur chaque registre. Ignore une colonne si elle n'existe pas dans le
    rapport ou si personne n'a de valeur positive dessus (évite d'afficher un 0 comme si
    c'était une menace réelle)."""
    threats = []
    for column, label, unit in PLAYER_THREAT_COLUMNS:
        best_name, best_value = None, None
        for p in players:
            v = p.get(column)
            if isinstance(v, (int, float)) and v > 0 and (best_value is None or v > best_value):
                best_name, best_value = p.get("Nom du joueur"), v
        if best_name:
            threats.append({"label": label, "name": best_name, "value": _num(best_value), "unit": unit})
    return threats


def compute_team_position_history(history, team_name):
    """history : liste de rapports Pro D2 successifs (du plus ancien au plus récent), sous
    la forme [{"label": "J1", "data": <rapport parsé>}, ...]. Renvoie l'évolution de la
    position au classement de l'équipe au fil des rapports importés — utile pour voir si
    une équipe est en forme ou en méforme, ce qu'un instantané seul ne montre pas. Se
    remplit progressivement semaine après semaine : vide ou à un seul point tant qu'un
    historique suffisant n'a pas été accumulé (un nouveau rapport par semaine)."""
    points = []
    for entry in history:
        classement = entry["data"].get("classement", [])
        for rank, row in enumerate(classement, start=1):
            if row.get("Nom de l'équipe") == team_name:
                points.append({"label": entry["label"], "rank": rank, "points": _num(row.get("Points"))})
                break
    return points


def compute_most_used_players(players):
    """Classe l'effectif par utilisation : d'abord par nombre de titularisations, puis par
    temps de jeu total, du plus utilisé au moins utilisé. Pas de poste assigné (l'export
    Pro D2 ne précise pas à quel poste chaque joueur a joué)."""
    rows = []
    for p in players:
        rows.append({
            "name": p.get("Nom du joueur"),
            "titulaires": p.get("Titulaires") or 0,
            "remplacant": p.get("Remplaçants") or 0,
            "matchs": p.get("Matchs") or 0,
            "temps_de_jeu": p.get("Temps de Jeu") or 0,
        })
    rows.sort(key=lambda r: (-(r["titulaires"] or 0), -(r["temps_de_jeu"] or 0)))
    return rows

def compute_player_groups(players):
    """Découpe les colonnes des joueurs d'une équipe en petits tableaux thématiques
    (même principe que les pages Attaque/Défense/Ruck des joueurs sur le reste du site)."""
    if not players:
        return []
    all_columns = list(players[0].keys())
    grouped_columns = {c for cols in PLAYER_COLUMN_GROUPS.values() for c in cols}
    leftover = [c for c in all_columns if c not in grouped_columns and c != "Nom du joueur"]

    groups = []
    for group_name, columns in PLAYER_COLUMN_GROUPS.items():
        cols_present = [c for c in columns if c in all_columns]
        if group_name == "Général" and leftover:
            cols_present = cols_present + leftover
        if not cols_present:
            continue
        groups.append({
            "name": group_name,
            "columns": cols_present,
            "rows": [{"name": p.get("Nom du joueur"), "values": [p.get(c) for c in cols_present]} for p in players],
        })
    return groups
